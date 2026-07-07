"""Excel workbook for the milestone impact diagnostic (backlog A3).

Consumes the plain-dict serializations — ``ImpactAnalysis.to_dict()`` (see
``scheduleiq.analytics.impact``) and ``AsBuiltReconstruction.to_dict()`` (see
``scheduleiq.analytics.asbuilt``) — and never imports the analytics modules
directly, so the caller controls what enrichment (e.g. a ``data_date`` key)
reaches the sheets.  Styling reuses the LI helpers from report/excel.py (teal
#1F6F7B headers, gray grid) so this workbook sits alongside the results,
trend, path-analysis, and intake-review workbooks without a visual seam;
excel.py itself is never modified.

Every sheet is stamped PRELIMINARY / ADR-0007-diagnostic in its title row —
these are engine-computed diagnostic deltas, not a competing schedule; the
tool-of-record dates remain the schedule (ADR-0007 §4, presentation rule).
"""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from .excel import STATUS_FILL, TEAL, _header, _row

PRELIM = ("PRELIMINARY — ADR-0007 engine diagnostic.  Tool-of-record dates "
         "remain the schedule; every figure below is a labelled delta, not a "
         "competing schedule.")


def _fmt(x, nd=1):
    if x is None:
        return ""
    if isinstance(x, bool):
        return "yes" if x else "no"
    if isinstance(x, float):
        return round(x, nd)
    return x


def _title(ws, text: str) -> int:
    """Write the sheet title (row 1) plus the PRELIMINARY stamp (row 2);
    returns the next free row."""
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    ws["A2"] = PRELIM
    ws["A2"].font = Font(italic=True, size=9, color="C00000")
    return 4


def _meta(ws, r, pairs):
    for k, v in pairs:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=v).font = Font(size=10)
        r += 1
    return r


def _notes(d: dict[str, Any]) -> str:
    """A compact, human-readable rendering of a scenario's scalar detail
    fields (list/dict-valued details get their own dedicated table/columns
    elsewhere, so they are skipped here)."""
    parts = []
    if d.get("blocking"):
        parts.append(f"BLOCKED: {d['blocking']}")
    for k, v in (d.get("details") or {}).items():
        if isinstance(v, (list, dict)):
            continue
        parts.append(f"{k}: {v}")
    return "; ".join(parts)


# ---------------------------------------------------------------------------
# Waterfall
# ---------------------------------------------------------------------------
def _write_waterfall(wb, impact: dict[str, Any]):
    ws = wb.active
    ws.title = "Waterfall"
    tgt = impact.get("target") or {}
    baseline = impact.get("baseline") or {}
    hs = impact.get("handshake") or {}
    r = _title(ws, f"Milestone Impact Diagnostic — {tgt.get('code') or tgt.get('uid') or '—'}")

    r = _meta(ws, r, [
        ("Target code", tgt.get("code")),
        ("Target name", tgt.get("name")),
        ("Target calendar", tgt.get("calendar")),
        ("Target resolved how", tgt.get("resolved_how")),
        ("Data date", impact.get("data_date")),
    ])
    r += 1
    ws.cell(row=r, column=1, value="Baseline — record vs. engine").font = Font(bold=True, size=11)
    r += 1
    r = _meta(ws, r, [
        ("Record early start", baseline.get("record_early_start")),
        ("Record early finish", baseline.get("record_early_finish")),
        ("Engine early start (diagnostic)", baseline.get("engine_early_start")),
        ("Engine early finish (diagnostic)", baseline.get("engine_early_finish")),
        ("Engine total float (wd)", baseline.get("engine_total_float_workdays")),
        ("Baseline computable", _fmt(baseline.get("computable"))),
    ])
    r += 1
    ws.cell(row=r, column=1, value="Validation handshake (SET-02)").font = Font(bold=True, size=11)
    r += 1
    r = _meta(ws, r, [
        ("Match rate (%)", hs.get("match_rate_pct")),
        ("Threshold (%)", hs.get("threshold_pct")),
        ("Passed", _fmt(hs.get("passed"))),
        ("Convention", hs.get("convention")),
        ("Lag strategy", hs.get("lag_strategy")),
        ("Statusing mode", hs.get("statusing_mode")),
        ("Constraint applications", hs.get("constraint_applications")),
    ])

    hrow = r + 2
    _header(ws, hrow, ["Scenario", "Delta (wd)", "Delta (cal-days)", "Computable",
                       "Target Finish (engine)", "Notes / Details"],
            [32, 12, 14, 12, 20, 80])
    rr = hrow + 1
    for d in impact.get("waterfall") or []:
        _row(ws, rr, [d.get("scenario"), d.get("delta_workdays"),
                     d.get("delta_calendar_days"), _fmt(d.get("computable")),
                     d.get("target_finish_engine"), _notes(d)],
             fill=None if d.get("computable") else STATUS_FILL.get("N/A"))
        rr += 1
    ws.auto_filter.ref = f"A{hrow}:F{max(rr - 1, hrow)}"


# ---------------------------------------------------------------------------
# Constraint Attribution
# ---------------------------------------------------------------------------
def _write_constraint_attribution(wb, impact: dict[str, Any]):
    ws = wb.create_sheet("Constraint Attribution")
    r = _title(ws, "Per-Constraint Attribution (waterfall detail)")
    _header(ws, r, ["Constraint (scenario)", "On Controlling Path", "Delta (wd)",
                    "Delta (cal-days)", "Computable", "Target Finish (engine)", "Notes"],
            [40, 16, 12, 14, 12, 20, 60])
    rr = r + 1
    for d in impact.get("constraint_attribution") or []:
        on_path = (d.get("details") or {}).get("on_controlling_path")
        _row(ws, rr, [d.get("scenario"), _fmt(on_path), d.get("delta_workdays"),
                     d.get("delta_calendar_days"), _fmt(d.get("computable")),
                     d.get("target_finish_engine"), _notes(d)],
             fill="C6E0B4" if on_path else None)
        rr += 1

    rr += 2
    ws.cell(row=rr, column=1, value="Float Absorbed (constraints released)").font = \
        Font(bold=True, size=11)
    rr += 1
    _header(ws, rr, ["Activity", "TF Constrained (wd)", "TF Unconstrained (wd)",
                     "Float Absorbed (wd)"], [14, 20, 22, 18])
    rr += 1
    all_off = next((d for d in impact.get("waterfall") or []
                    if d.get("scenario") == "constraints_released_all"), None)
    for row in ((all_off or {}).get("details") or {}).get("float_absorbed_workdays", []) or []:
        _row(ws, rr, [row.get("code"), row.get("tf_constrained_workdays"),
                     row.get("tf_unconstrained_workdays"),
                     row.get("float_absorbed_workdays")])
        rr += 1


# ---------------------------------------------------------------------------
# Criticality (P5)
# ---------------------------------------------------------------------------
def _write_criticality(wb, impact: dict[str, Any]):
    ws = wb.create_sheet("Criticality (P5)")
    r = _title(ws, "Constraint-Free Criticality (P5)")
    blk = impact.get("constraint_free_criticality") or {}
    if blk.get("reason"):
        ws.cell(row=r, column=1, value=blk["reason"]).font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [
        ("Criticality definition", blk.get("criticality_definition")),
        ("Incomplete activities", blk.get("incomplete_activities")),
        ("% critical with constraints", blk.get("pct_critical_with_constraints")),
        ("% critical without constraints", blk.get("pct_critical_without")),
        ("# manufactured-critical", blk.get("n_manufactured")),
        ("# masked-critical", blk.get("n_masked")),
        ("Target TF with constraints (wd)",
         (blk.get("target_total_float_workdays") or {}).get("with_constraints")),
        ("Target TF without constraints (wd)",
         (blk.get("target_total_float_workdays") or {}).get("without")),
    ])
    hrow = r + 2
    _header(ws, hrow, ["Activity", "Critical With Constraints", "Critical Without",
                       "Class"], [14, 22, 18, 18])
    rr = hrow + 1
    for row in blk.get("flip_table") or []:
        cb, cu = row.get("critical_with_constraints"), row.get("critical_without")
        cls = "manufactured" if (cb and not cu) else ("masked" if (cu and not cb) else "stable")
        _row(ws, rr, [row.get("code"), _fmt(cb), _fmt(cu), cls],
             fill=STATUS_FILL.get("WARNING") if cls == "manufactured" else
             (STATUS_FILL.get("INFO") if cls == "masked" else None))
        rr += 1


# ---------------------------------------------------------------------------
# Calendar Restatement
# ---------------------------------------------------------------------------
def _write_calendar_restatement(wb, impact: dict[str, Any]):
    ws = wb.create_sheet("Calendar Restatement")
    r = _title(ws, "Calendar-Neutral Restatement (controlling path)")
    blk = impact.get("calendar_restatement") or {}
    if blk.get("reason"):
        ws.cell(row=r, column=1, value=blk["reason"]).font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [
        ("Default (project) calendar", blk.get("default_calendar")),
        ("Default hours/day", blk.get("default_hours_per_day")),
        ("Max divergence on path (days)", blk.get("max_divergence_days_on_path")),
        ("Target divergence (days)", blk.get("target_divergence_days")),
        ("Note", blk.get("note")),
    ])
    hrow = r + 2
    _header(ws, hrow, ["Activity", "Own hrs/day", "Total Float (h)",
                       "Day-Float (own cal.)", "Day-Float (default cal.)",
                       "Divergence (days)", "Diverges"],
            [14, 12, 16, 20, 22, 18, 12])
    rr = hrow + 1
    for row in blk.get("diverging_activities") or []:
        _row(ws, rr, [row.get("code"), row.get("own_hours_per_day"),
                     row.get("total_float_hours"), row.get("day_float_own_calendar"),
                     row.get("day_float_default_calendar"), row.get("divergence_days"),
                     _fmt(row.get("diverges"))],
             fill=STATUS_FILL.get("WARNING") if row.get("diverges") else None)
        rr += 1


# ---------------------------------------------------------------------------
# Open Ends
# ---------------------------------------------------------------------------
def _write_open_ends(wb, impact: dict[str, Any]):
    ws = wb.create_sheet("Open Ends")
    r = _title(ws, "Open Ends — No Forward Path to the Target")
    blk = impact.get("open_ends") or {}
    r = _meta(ws, r, [
        ("Count", blk.get("count")),
        ("Truncated", _fmt(blk.get("truncated"))),
        ("Note", blk.get("note")),
    ])
    hrow = r + 2
    _header(ws, hrow, ["Activity", "Name", "Status"], [14, 40, 16])
    rr = hrow + 1
    for a in blk.get("activities") or []:
        _row(ws, rr, [a.get("code"), a.get("name"), a.get("status")])
        rr += 1


# ---------------------------------------------------------------------------
# As-Built Paths (P6)
# ---------------------------------------------------------------------------
def _write_asbuilt(wb, asbuilt: dict[str, Any]):
    ws = wb.create_sheet("As-Built Paths (P6)")
    r = _title(ws, "As-Built Path Reconstruction (P6)")
    summary = asbuilt.get("summary") or {}
    r = _meta(ws, r, [
        ("Label", asbuilt.get("label")),
        ("End anchor", asbuilt.get("end_anchor_code")),
        ("End anchor resolution", asbuilt.get("end_anchor_resolution")),
        ("Lag strategy", asbuilt.get("lag_strategy")),
        ("Gap threshold (wd)", asbuilt.get("gap_threshold_wd")),
        ("Top N", asbuilt.get("top_n")),
        ("Span convention", asbuilt.get("span_convention")),
        ("Reason", asbuilt.get("reason") or "—"),
        ("Actualized activities", summary.get("actualized_activities")),
        ("Actualized relationships", summary.get("actualized_relationships")),
        ("Contradicted relationships", summary.get("contradicted_relationships")),
        ("Unreached started activities", summary.get("unreached_started_activities")),
    ])

    hrow = r + 2
    ws.cell(row=hrow - 1, column=1, value="Chain links").font = Font(bold=True, size=11)
    _header(ws, hrow, ["Chain Rank", "Pred", "Succ", "Rel Type", "Planned Lag (wd)",
                       "Actual Lag (wd)", "Tightness (wd)", "Pred Anchor Date",
                       "Succ Anchor Date", "Lag Calendar", "Calendar Fallback",
                       "Is Gap"],
            [10, 12, 12, 10, 14, 14, 14, 16, 16, 20, 14, 10])
    rr = hrow + 1
    for chain in asbuilt.get("chains") or []:
        for lk in chain.get("links") or []:
            _row(ws, rr, [chain.get("rank"), lk.get("pred_code"), lk.get("succ_code"),
                         lk.get("rel_type"), lk.get("planned_lag_wd"),
                         lk.get("actual_lag_wd"), lk.get("tightness_wd"),
                         lk.get("pred_anchor_date"), lk.get("succ_anchor_date"),
                         lk.get("lag_calendar"), _fmt(lk.get("lag_calendar_fallback")),
                         _fmt(lk.get("is_gap"))],
                 fill=STATUS_FILL.get("WARNING") if lk.get("is_gap") else None)
            rr += 1

    rr += 2
    ws.cell(row=rr, column=1, value="Contradicted (out-of-sequence) logic"
            ).font = Font(bold=True, size=11)
    rr += 1
    _header(ws, rr, ["Pred", "Succ", "Rel Type", "Planned Lag (wd)", "Actual Lag (wd)",
                     "Tightness (wd)", "Pred Anchor Date", "Succ Anchor Date"],
            [12, 12, 10, 14, 14, 14, 16, 16])
    rr += 1
    for lk in asbuilt.get("contradicted_links") or []:
        _row(ws, rr, [lk.get("pred_code"), lk.get("succ_code"), lk.get("rel_type"),
                     lk.get("planned_lag_wd"), lk.get("actual_lag_wd"),
                     lk.get("tightness_wd"), lk.get("pred_anchor_date"),
                     lk.get("succ_anchor_date")],
             fill=STATUS_FILL.get("FAIL"))
        rr += 1


# ---------------------------------------------------------------------------
# Disclosures
# ---------------------------------------------------------------------------
def _write_disclosures(wb, impact: dict[str, Any], asbuilt: dict[str, Any]):
    ws = wb.create_sheet("Disclosures")
    r = _title(ws, "Disclosures and Deferred Items")

    def _block(title, items):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11)
        r += 1
        if not items:
            ws.cell(row=r, column=1, value="— none —").font = Font(italic=True, size=10)
            r += 1
        for it in items:
            ws.cell(row=r, column=1, value=str(it)).font = Font(size=10)
            r += 1
        r += 1

    _block("Impact analytics disclosures", impact.get("disclosures") or [])
    _block("Impact analytics — deferred", impact.get("deferred") or [])
    _block("As-built reconstruction disclosures", asbuilt.get("disclosures") or [])
    ws.column_dimensions["A"].width = 110


def write_impact_workbook(impact_dict: dict[str, Any], asbuilt_dict: dict[str, Any],
                          out_path: str) -> str:
    """Write the milestone-impact-diagnostic workbook: Waterfall, Constraint
    Attribution, Criticality (P5), Calendar Restatement, Open Ends, As-Built
    Paths (P6), and Disclosures.  ``impact_dict``/``asbuilt_dict`` are the
    plain-dict serializations from ``ImpactAnalysis.to_dict()`` /
    ``AsBuiltReconstruction.to_dict()`` (optionally enriched, e.g. with a
    ``data_date`` key on ``impact_dict``)."""
    wb = Workbook()
    _write_waterfall(wb, impact_dict)
    _write_constraint_attribution(wb, impact_dict)
    _write_criticality(wb, impact_dict)
    _write_calendar_restatement(wb, impact_dict)
    _write_open_ends(wb, impact_dict)
    _write_asbuilt(wb, asbuilt_dict)
    _write_disclosures(wb, impact_dict, asbuilt_dict)
    wb.save(out_path)
    return out_path
