"""Tests for the ScheduleIQ v0.4 wiring wave — connecting every module built
this version (TIA workbench, weather overlay, work-pattern/edit-session
forensics, Ribbon/Phase/Compliance analyzers, the interactive cockpit,
benchmark-corpus context, damages/exposure, and the privileged internal
workbook) into ``runner.run`` and the Word report.

Exercises ``runner.load_run_config`` (the reserved ``config:`` profile
extension), the additive runner.py blocks, ``report/excel_v04.py``'s two
workbook writers, and the new "SCHEDULE FORENSICS SUPPLEMENT" report section
— using the demo_hs1.xer/demo_hs2.xer pair (both handshake at 100%; see
tests/fixtures/make_fixtures.py) plus the already-checked-in weather station
and TIA event CSV fixtures.
"""
import os
import subprocess
import sys

import openpyxl
import pytest
import yaml

SRC = os.path.join(os.path.dirname(__file__), "..", "src")
sys.path.insert(0, os.path.abspath(SRC))

from scheduleiq.cpm.handshake import clear_handshake_cache            # noqa: E402
from scheduleiq.metrics.engine import load_profile                    # noqa: E402
import scheduleiq.runner as runner_mod                                # noqa: E402
from scheduleiq.runner import load_run_config, run                    # noqa: E402

FIX = os.path.join(os.path.dirname(__file__), "fixtures")
HS1 = os.path.join(FIX, "demo_hs1.xer")
HS2 = os.path.join(FIX, "demo_hs2.xer")
BASELINE = os.path.join(FIX, "demo_baseline.xer")
U1 = os.path.join(FIX, "demo_update1.xer")
U2 = os.path.join(FIX, "demo_update2.xer")
WEATHER_CSV = os.path.join(FIX, "weather_station_sample.csv")
TIA_EVENTS_CSV = os.path.join(FIX, "tia_events_sample.csv")
LEGACY_PROFILE = os.path.join(os.path.dirname(__file__), "..", "profiles",
                              "example_profile.yaml")

_SRA_TEST_ITERATIONS = 50

V04_SHEETS_MIN = {"Summary", "Ribbon", "Phase", "Disclosures"}
V04_SHEETS_FULL = V04_SHEETS_MIN | {"TIA Impacts", "Collapse", "Weather Exhibit",
                                    "Work Patterns", "Edit Sessions", "Compliance"}
INTERNAL_SHEETS = {"LI-11 SMI", "LI-12 DDI", "LI-13 ARR", "LI-14 PPS", "LI-15 RSA",
                   "Robustness Certificate"}


@pytest.fixture(scope="session", autouse=True)
def fixtures():
    if not (os.path.exists(HS1) and os.path.exists(HS2)):
        subprocess.run([sys.executable, os.path.join(FIX, "make_fixtures.py")],
                       check=True)


@pytest.fixture(autouse=True)
def _fresh_handshake_cache():
    clear_handshake_cache()
    yield
    clear_handshake_cache()


def _write_full_config(tmp_path, corpus_path, *, record: bool) -> str:
    cfg = {
        "DCMA-03": 10,
        "config": {
            "damages": {
                "ld_rate_per_day": 25000,
                "contractual_completion": "2025-08-01",
                "time_cost_per_day": 12000,
                "currency": "USD",
            },
            "weather": {
                "station_csv": WEATHER_CSV,
                "sensitive_tags": ["foundation", "steel", "groundwork"],
            },
            "corpus": {
                "path": str(corpus_path),
                "sector": "process-plant",
                "record": record,
            },
            "tia_events": TIA_EVENTS_CSV,
            "cockpit": True,
            "internal_workbook": True,
        },
    }
    p = tmp_path / "full_config.yaml"
    p.write_text(yaml.safe_dump(cfg, sort_keys=False))
    return str(p)


def _all_cell_values(ws) -> list:
    vals = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                vals.append(cell.value)
    return vals


# --------------------------------------------------------------- load_run_config
def test_load_run_config_empty_for_flat_profile():
    assert load_run_config(LEGACY_PROFILE) == {}
    # and the flat threshold path is completely unaffected by the presence of
    # this function / the (absent) config: key
    assert load_profile(LEGACY_PROFILE) == {"DCMA-03": 10.0, "DCMA-06": 10.0,
                                            "DAT-04": 30.0}


def test_load_run_config_none_and_json_are_empty(tmp_path):
    assert load_run_config(None) == {}
    j = tmp_path / "flat.json"
    j.write_text('{"DCMA-03": 5}')
    assert load_run_config(str(j)) == {}


def test_load_run_config_reads_reserved_key(tmp_path):
    p = tmp_path / "cfg.yaml"
    p.write_text("DCMA-03: 10\nconfig:\n  cockpit: false\n  internal_workbook: true\n")
    cfg = load_run_config(str(p))
    assert cfg == {"cockpit": False, "internal_workbook": True}
    # the flat override path still parses the threshold line untouched
    assert load_profile(str(p))["DCMA-03"] == 10.0


def test_flat_only_profile_config_is_empty_regression():
    """A profile with ONLY flat 'ID: threshold' lines (no config: section) —
    the v0.4 extension must not invent config out of nothing."""
    assert load_run_config(LEGACY_PROFILE) == {}


# --------------------------------------------------------- full-config e2e run
def test_runner_full_config_hs_pair(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    corpus_path = tmp_path / "corpus.jsonl"
    cfg = _write_full_config(tmp_path, corpus_path, record=False)
    out = str(tmp_path / "out")

    rr = run([HS1, HS2], out, profile=cfg, make_pdf=False)

    # messages clean of errors — "skipped:"/"error:" mark an actual exception
    # in a block; a fresh (empty) corpus legitimately REFUSES percentile
    # placement (small-n honesty, § 6.9) and that disclosure is expected, not
    # a failure, so it is excluded from the error scan.
    bad = [m for m in rr.messages
          if ("skipped:" in m or " error" in m.lower()
              or "handshake below threshold" in m)
          and "percentile placement refused" not in m]
    assert not bad, rr.messages

    expected = ["v04_analytics_supplement.xlsx", "cockpit.html",
               "INTERNAL_PRIVILEGED_workbook.xlsx", "schedule_assessment.docx"]
    for name in expected:
        p = os.path.join(out, name)
        assert os.path.exists(p) and os.path.getsize(p) > 300, f"missing/empty {name}"
        assert p in rr.outputs

    # v0.4 workbook: full sheet set, exposure column present on TIA Impacts
    wb = openpyxl.load_workbook(os.path.join(out, "v04_analytics_supplement.xlsx"))
    assert V04_SHEETS_FULL <= set(wb.sheetnames)
    ws = wb["TIA Impacts"]
    header_row = [c.value for c in next(
        r for r in ws.iter_rows() if r[0].value == "Event")]
    assert "Exposure (cumulative, time-cost)" in header_row
    ws_collapse = wb["Collapse"]
    collapse_header = [c.value for c in next(
        r for r in ws_collapse.iter_rows() if r[0].value == "Party")]
    assert "Exposure (time-cost)" in collapse_header

    # internal privileged workbook: expected sheets, privileged banner on each
    wb2 = openpyxl.load_workbook(os.path.join(out, "INTERNAL_PRIVILEGED_workbook.xlsx"))
    assert INTERNAL_SHEETS <= set(wb2.sheetnames)
    for name in wb2.sheetnames:
        found = any(cell.value and "PRIVILEGED" in str(cell.value)
                   for row in wb2[name].iter_rows(max_row=4) for cell in row)
        assert found, f"sheet {name!r} missing a PRIVILEGED banner"

    # corpus file NOT written (record: false)
    assert not os.path.exists(corpus_path)
    assert any("NOT recorded" in m for m in rr.messages)


def test_runner_full_config_corpus_record_true_writes_one_row(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    corpus_path = tmp_path / "corpus.jsonl"
    cfg = _write_full_config(tmp_path, corpus_path, record=True)
    out = str(tmp_path / "out")

    rr = run([HS1, HS2], out, profile=cfg, make_pdf=False)
    assert any("corpus WRITE" in m for m in rr.messages), rr.messages
    assert os.path.exists(corpus_path)
    with open(corpus_path, encoding="utf-8") as f:
        lines = [ln for ln in f.read().splitlines() if ln.strip()]
    assert len(lines) == 1


# --------------------------------------------------------------- CLI overrides
def test_cli_no_cockpit_and_internal_workbook_flags(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    out = str(tmp_path / "out")
    rr = run([HS1, HS2], out, make_pdf=False, no_cockpit=True, internal_workbook=True)
    assert not os.path.exists(os.path.join(out, "cockpit.html"))
    assert os.path.exists(os.path.join(out, "INTERNAL_PRIVILEGED_workbook.xlsx"))


# --------------------------------------------------------- legacy regressions
def test_runner_legacy_flat_profile_no_extra_v04_artifacts(tmp_path, monkeypatch):
    """profiles/example_profile.yaml carries ONLY flat threshold overrides —
    behavior must be identical to before: no damages/TIA/weather/corpus/
    internal-workbook artifacts, only the default-on v0.4 additions (cockpit
    + the Ribbon/Phase-only v0.4 workbook), and no crash."""
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    out = str(tmp_path / "out")
    rr = run([HS1, HS2], out, profile=LEGACY_PROFILE, make_pdf=False)

    assert not rr.messages, rr.messages
    assert not os.path.exists(os.path.join(out, "INTERNAL_PRIVILEGED_workbook.xlsx"))
    assert os.path.exists(os.path.join(out, "cockpit.html"))

    wb = openpyxl.load_workbook(os.path.join(out, "v04_analytics_supplement.xlsx"))
    sheets = set(wb.sheetnames)
    assert V04_SHEETS_MIN <= sheets
    # a work-pattern/TIA/weather-free run: none of those sheets appear
    assert not ({"TIA Impacts", "Collapse", "Weather Exhibit"} & sheets)
    # the hs pair is a 2-schedule series: work patterns/edit sessions/
    # compliance DO run (they need no config), so they ARE present
    assert {"Work Patterns", "Edit Sessions", "Compliance"} <= sheets


def test_runner_legacy_3file_series_refusals_preserved(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    out = str(tmp_path / "out")
    rr = run([BASELINE, U1, U2], out, make_pdf=False)

    assert any("forensic delay diagnostics refused" in m for m in rr.messages), rr.messages
    assert any("schedule risk analysis refused" in m for m in rr.messages), rr.messages

    # record-only v0.4 analytics don't need the engine — still produced
    assert os.path.exists(os.path.join(out, "cockpit.html"))
    wb = openpyxl.load_workbook(os.path.join(out, "v04_analytics_supplement.xlsx"))
    sheets = set(wb.sheetnames)
    assert {"Ribbon", "Phase", "Compliance"} <= sheets
    # no engine-backed TIA/collapse content leaked through
    assert "TIA Impacts" not in sheets


def test_runner_single_file_v04_silent_and_ribbon_phase_only(tmp_path, monkeypatch):
    """A single-file run: work patterns/edit sessions/compliance/TIA all need
    >= 2 schedules and must no-op SILENTLY (no message), matching the
    forensic-diagnostics block's established convention; Ribbon/Phase still
    run (single-file capable)."""
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    out = str(tmp_path / "out")
    rr = run([HS2], out, make_pdf=False)
    assert not any("work-pattern" in m or "editing-session" in m
                  or "compliance" in m or "TIA" in m for m in rr.messages), rr.messages

    wb = openpyxl.load_workbook(os.path.join(out, "v04_analytics_supplement.xlsx"))
    sheets = set(wb.sheetnames)
    assert {"Ribbon", "Phase"} <= sheets
    assert not ({"Work Patterns", "Edit Sessions", "Compliance", "TIA Impacts"} & sheets)


# ------------------------------------------------------------------- word report
def test_word_report_supplement_section_no_provocative_leak(tmp_path, monkeypatch):
    import re
    import zipfile

    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    corpus_path = tmp_path / "corpus.jsonl"
    cfg = _write_full_config(tmp_path, corpus_path, record=False)
    out = str(tmp_path / "out")
    rr = run([HS1, HS2], out, profile=cfg, make_pdf=False)

    docx = os.path.join(out, "schedule_assessment.docx")
    assert docx in rr.outputs
    with zipfile.ZipFile(docx) as z:
        doc = z.read("word/document.xml").decode("utf-8")

    assert "SCHEDULE FORENSICS SUPPLEMENT" in doc
    assert "Weather-Delay Exhibit" in doc
    assert "Ribbon Analysis" in doc

    # LI-11..LI-15 provocative indices are PRIVILEGED — never in the report
    leak_patterns = [r"\bLI-1[1-5]\b", r"\bSMI\b", r"\bDDI\b", r"\bARR\b",
                     r"\bPPS\b", r"\bRSA\b", "Manipulation Indicator",
                     "Directed Date Index", "Attribution Robustness",
                     "Pacing Plausibility", "Rebuttal Surface"]
    for pat in leak_patterns:
        assert not re.search(pat, doc), f"provocative-index leak into report: {pat!r}"


def test_report_builder_never_imports_li_provocative():
    """report_builder.py may DISCUSS in comments/docstrings why the
    provocative indices are excluded (that string literally names the
    module it must not touch), but it must never actually import it."""
    import scheduleiq.report.report_builder as rb
    src = open(rb.__file__, encoding="utf-8").read()
    for line in src.splitlines():
        stripped = line.strip()
        if stripped.startswith(("import ", "from ")) and "li_provocative" in stripped:
            pytest.fail(f"report_builder.py imports li_provocative: {stripped!r}")


# ---------------------------------------------------------------------- determinism
def test_v04_workbook_deterministic_on_reread(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    corpus_path = tmp_path / "corpus.jsonl"
    cfg = _write_full_config(tmp_path, corpus_path, record=False)
    out1 = str(tmp_path / "run1")
    out2 = str(tmp_path / "run2")
    run([HS1, HS2], out1, profile=cfg, make_pdf=False)
    run([HS1, HS2], out2, profile=cfg, make_pdf=False)

    wb1 = openpyxl.load_workbook(os.path.join(out1, "v04_analytics_supplement.xlsx"))
    wb2 = openpyxl.load_workbook(os.path.join(out2, "v04_analytics_supplement.xlsx"))
    assert wb1.sheetnames == wb2.sheetnames
    for name in wb1.sheetnames:
        v1 = _all_cell_values(wb1[name])
        v2 = _all_cell_values(wb2[name])
        assert v1 == v2, f"v0.4 sheet {name!r} not deterministic between two runs"
