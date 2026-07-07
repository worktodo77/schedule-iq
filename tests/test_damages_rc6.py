"""Tests for the damages/exposure overlay (backlog S7, ANALYTICS_PROPOSAL.md
§6.6) and the RC6 public-spec publication build (scripts/build_public_spec.py).

Exposure arithmetic: hand-computed LD and time-cost lines (exact, including
the "14 cd x $25,000/cd = $350,000" example from the design doc), calendar
vs. workday basis, and the always-present standing label.  None-config
byte-identity for one workbook (excel_impact) and one figure (waterfall) —
the hard regression gate: a None damages config must not move a single byte
of pre-existing output.  With-config workbooks/figures must gain the new
Exposure surface and remain deterministic across repeated writes.

RC6: runs scripts/build_public_spec.py against a temp copy of the repo and
asserts the published package is well-formed — scorecard.yaml copied
verbatim modulo the internal_variant strip, LICENSE present, the spec
stamped, the sample CSV present, and reference_scorer.py reproducing the
expected score exactly.
"""
import importlib.util
import os
import shutil
import subprocess
import sys

import openpyxl
import pytest
import yaml

SRC = os.path.join(os.path.dirname(__file__), "..", "src")
sys.path.insert(0, os.path.abspath(SRC))

from scheduleiq.analytics.damages import (DamagesConfig, STANDING_LABEL,        # noqa: E402
                                          exposure_for_date, exposure_for_delta,
                                          load_damages_config)
from scheduleiq.ingest import load                                              # noqa: E402
from scheduleiq.analytics.impact import run_impact_analysis                     # noqa: E402
from scheduleiq.analytics.asbuilt import reconstruct_asbuilt_paths              # noqa: E402
from scheduleiq.report.excel_impact import write_impact_workbook                # noqa: E402
from scheduleiq.report.impact_figures import waterfall_figure                   # noqa: E402

ROOT = os.path.join(os.path.dirname(__file__), "..")
FIX = os.path.join(os.path.dirname(__file__), "fixtures")
IMPACT = os.path.join(FIX, "demo_impact.xer")
BUILD_SCRIPT = os.path.join(ROOT, "scripts", "build_public_spec.py")


@pytest.fixture(scope="session", autouse=True)
def fixtures():
    if not os.path.exists(IMPACT):
        subprocess.run([sys.executable, os.path.join(FIX, "make_fixtures.py")],
                       check=True)


# ===========================================================================
# 1. Exposure arithmetic (analytics/damages.py)
# ===========================================================================
def test_exposure_for_delta_matches_design_doc_example():
    """ANALYTICS_PROPOSAL.md §6.6's worked example: "14 cd × $25,000/cd =
    $350,000"."""
    cfg = DamagesConfig(time_cost_per_day=25000.0)
    line = exposure_for_delta(14, cfg, basis_note="calendar days")
    assert line.amount == 350000.0
    assert line.formula_text == "14 cd × $25,000/cd = $350,000"
    assert line.label == STANDING_LABEL


def test_exposure_for_delta_hand_computed_negative():
    cfg = DamagesConfig(time_cost_per_day=1200.0)
    line = exposure_for_delta(-9, cfg, basis_note="workdays (target calendar)")
    assert line.amount == -9 * 1200.0 == -10800.0
    assert line.formula_text == "-9 wd × $1,200/wd = -$10,800"


def test_exposure_for_delta_none_without_rate():
    assert exposure_for_delta(14, None).amount is None
    assert exposure_for_delta(14, DamagesConfig()).amount is None
    assert exposure_for_delta(None, DamagesConfig(time_cost_per_day=100)).amount is None


def test_exposure_for_date_hand_computed_ld():
    cfg = DamagesConfig(ld_rate_per_day=25000.0, contractual_completion="2026-01-01",
                        time_cost_per_day=1.0)
    line = exposure_for_date("2026-01-15", cfg)
    assert line.amount == 14 * 25000.0 == 350000.0
    assert line.formula_text == "14 cd × $25,000/cd = $350,000"


def test_exposure_for_date_clamped_at_zero_when_early():
    cfg = DamagesConfig(ld_rate_per_day=25000.0, contractual_completion="2026-01-01",
                        time_cost_per_day=1.0)
    line = exposure_for_date("2025-12-20", cfg)
    assert line.amount == 0.0
    assert "AHEAD" in line.formula_text


def test_exposure_for_date_disabled_without_both_rate_and_date():
    assert exposure_for_date("2026-01-01", None).amount is None
    assert exposure_for_date("2026-01-01", DamagesConfig(ld_rate_per_day=100)).amount is None
    assert exposure_for_date(
        "2026-01-01", DamagesConfig(contractual_completion="2026-01-01")).amount is None


def test_exposure_for_date_is_always_calendar_days_regardless_of_daily_basis():
    """LD prices in calendar days even when daily_basis == 'workday' — fixed
    convention, documented, never left to the config."""
    cfg = DamagesConfig(ld_rate_per_day=100.0, contractual_completion="2026-01-01",
                        daily_basis="workday")
    line = exposure_for_date("2026-01-11", cfg)
    assert line.amount == 10 * 100.0  # 10 CALENDAR days, not workdays
    assert "calendar days" in line.basis


def test_exposure_for_delta_basis_mismatch_is_flagged_not_silently_converted():
    cfg = DamagesConfig(time_cost_per_day=100.0, daily_basis="calendar")
    line = exposure_for_delta(5, cfg, basis_note="workdays (target calendar)")
    # arithmetic is always literal -- 5 * 100, no unit conversion applied
    assert line.amount == 500.0
    assert "NOTE" in line.basis and "workday" in line.basis.lower()


def test_standing_label_present_on_every_exposure_line():
    cfg = DamagesConfig(ld_rate_per_day=1, contractual_completion="2026-01-01",
                        time_cost_per_day=1)
    for line in (exposure_for_delta(1, cfg), exposure_for_date("2026-01-02", cfg),
                exposure_for_delta(None, None), exposure_for_date(None, None)):
        assert line.label == STANDING_LABEL
        assert "quantum, causation, and entitlement are reserved to the expert" \
            in line.label


def test_load_damages_config_from_dict_and_none():
    assert load_damages_config(None) is None
    assert load_damages_config({}) is None
    cfg = load_damages_config({"ld_rate_per_day": 500, "time_cost_per_day": 200})
    assert isinstance(cfg, DamagesConfig)
    assert cfg.currency == "USD"
    assert cfg.daily_basis == "calendar"
    assert cfg.ld_enabled is False  # no contractual_completion


def test_load_damages_config_from_yaml_path(tmp_path):
    p = tmp_path / "damages.yaml"
    p.write_text("ld_rate_per_day: 30000\n"
                 "contractual_completion: '2026-03-01'\n"
                 "time_cost_per_day: 8000\n"
                 "currency: GBP\n"
                 "daily_basis: workday\n")
    cfg = load_damages_config(str(p))
    assert cfg.currency == "GBP"
    assert cfg.daily_basis == "workday"
    assert cfg.ld_enabled is True
    assert cfg.contractual_completion.isoformat() == "2026-03-01"


def test_load_damages_config_rejects_unknown_keys():
    with pytest.raises(ValueError):
        load_damages_config({"not_a_real_field": 1})


def test_damages_config_rejects_bad_daily_basis():
    with pytest.raises(ValueError):
        DamagesConfig(daily_basis="fortnight")


# ===========================================================================
# 2. Report surfaces — None-config byte-identity (hard regression gate)
# ===========================================================================
@pytest.fixture(scope="session")
def impact_dict():
    s = load(IMPACT)[0]
    return run_impact_analysis(s).to_dict()


@pytest.fixture(scope="session")
def asbuilt_dict():
    s = load(IMPACT)[0]
    return reconstruct_asbuilt_paths(s).to_dict()


DAMAGES_CFG = DamagesConfig(ld_rate_per_day=10000.0, contractual_completion="2024-01-01",
                            time_cost_per_day=5000.0)


def test_impact_workbook_none_config_byte_identical(impact_dict, asbuilt_dict, tmp_path):
    """A None damages config must reproduce byte-identical output to the
    pre-S7 workbook -- adding the parameter changes nothing by default."""
    p1 = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "a.xlsx"))
    p2 = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "b.xlsx"),
                               damages=None)
    with open(p1, "rb") as f1, open(p2, "rb") as f2:
        assert f1.read() == f2.read()


def test_waterfall_figure_none_config_byte_identical(impact_dict, tmp_path):
    p1 = waterfall_figure(impact_dict, str(tmp_path / "a.png"))
    p2 = waterfall_figure(impact_dict, str(tmp_path / "b.png"), damages=None)
    with open(p1, "rb") as f1, open(p2, "rb") as f2:
        assert f1.read() == f2.read()


# ===========================================================================
# 3. Report surfaces — with-config behavior
# ===========================================================================
def test_impact_workbook_with_damages_has_exposure_sheet_and_label(
        impact_dict, asbuilt_dict, tmp_path):
    path = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "d.xlsx"),
                                 damages=DAMAGES_CFG)
    wb = openpyxl.load_workbook(path)
    assert "Exposure" in wb.sheetnames
    ws = wb["Exposure"]
    all_vals = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert any(STANDING_LABEL in v for v in all_vals)
    assert any("cd ×" in v and "=" in v for v in all_vals), \
        "no priced formula line found on the Exposure sheet"
    # the Waterfall sheet gains the additive Exposure column
    wf = wb["Waterfall"]
    headers = []
    for row in wf.iter_rows(min_row=1, max_row=wf.max_row):
        vals = [c.value for c in row]
        if "Scenario" in vals:
            headers = vals
            break
    assert headers, "could not locate the Waterfall sheet's scenario header row"
    assert any(h and "Exposure" in str(h) for h in headers)


def test_impact_workbook_with_damages_deterministic(impact_dict, asbuilt_dict, tmp_path):
    p1 = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "a.xlsx"),
                               damages=DAMAGES_CFG)
    p2 = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "b.xlsx"),
                               damages=DAMAGES_CFG)
    with open(p1, "rb") as f1, open(p2, "rb") as f2:
        assert f1.read() == f2.read()


def test_waterfall_figure_with_damages_differs_and_is_deterministic(impact_dict, tmp_path):
    none_png = waterfall_figure(impact_dict, str(tmp_path / "none.png"))
    p1 = waterfall_figure(impact_dict, str(tmp_path / "a.png"), damages=DAMAGES_CFG)
    p2 = waterfall_figure(impact_dict, str(tmp_path / "b.png"), damages=DAMAGES_CFG)
    with open(p1, "rb") as f1, open(p2, "rb") as f2:
        assert f1.read() == f2.read(), "same config must render byte-identical figures"
    with open(none_png, "rb") as f0, open(p1, "rb") as f1:
        assert f0.read() != f1.read(), "a damages config must visibly change the figure"


# ===========================================================================
# 4. RC6 — scripts/build_public_spec.py
# ===========================================================================
@pytest.fixture(scope="module")
def built_public_spec_dir(tmp_path_factory):
    """Run the real build script against a TEMP COPY of the repo (so this
    test never mutates the checked-in docs/public_spec/ or scorecard.yaml)
    and return the path to the built public_spec directory."""
    work = tmp_path_factory.mktemp("rc6_build")
    repo_copy = work / "repo"
    shutil.copytree(ROOT, repo_copy, ignore=shutil.ignore_patterns(
        ".git", "__pycache__", "*.pyc", ".pytest_cache"))
    result = subprocess.run(
        [sys.executable, str(repo_copy / "scripts" / "build_public_spec.py")],
        capture_output=True, text=True, cwd=str(repo_copy))
    assert result.returncode == 0, f"build_public_spec.py failed:\n{result.stdout}\n{result.stderr}"
    return repo_copy / "docs" / "public_spec", result.stdout


def test_build_public_spec_runs_cleanly(built_public_spec_dir):
    _pub_dir, stdout = built_public_spec_dir
    assert "build complete" in stdout
    assert "Stripped internal_variant" in stdout


def test_build_public_spec_produces_expected_files(built_public_spec_dir):
    pub_dir, _stdout = built_public_spec_dir
    for name in ("scorecard.yaml", "LICENSE", "LI-RC-spec.md", "README.md",
                "sample_results.csv", "sample_results_expected_output.txt",
                "reference_scorer.py"):
        assert (pub_dir / name).exists(), f"missing {name}"


def test_build_public_spec_scorecard_verbatim_modulo_internal_variant_strip(
        built_public_spec_dir):
    pub_dir, _stdout = built_public_spec_dir
    src_path = os.path.join(ROOT, "src", "scheduleiq", "scorecard.yaml")
    with open(src_path, encoding="utf-8") as f:
        src_parsed = yaml.safe_load(f)
    with open(pub_dir / "scorecard.yaml", encoding="utf-8") as f:
        pub_text = f.read()
        pub_parsed = yaml.safe_load(pub_text)

    assert "internal_variant" in src_parsed, \
        "fixture assumption broken: source scorecard.yaml no longer has internal_variant"
    assert "internal_variant" not in pub_parsed, \
        "published scorecard.yaml must NOT carry the internal_variant block"
    assert "internal_variant:" not in pub_text

    # everything else is verbatim
    src_minus = {k: v for k, v in src_parsed.items() if k != "internal_variant"}
    assert pub_parsed == src_minus


def test_build_public_spec_license_is_apache_2_0_with_li_copyright(built_public_spec_dir):
    pub_dir, _stdout = built_public_spec_dir
    text = (pub_dir / "LICENSE").read_text(encoding="utf-8")
    assert "Apache License" in text and "Version 2.0" in text
    assert "Copyright 2026 Long International, Inc." in text


def test_build_public_spec_stamps_readme_and_spec_banners(built_public_spec_dir):
    pub_dir, _stdout = built_public_spec_dir
    readme = (pub_dir / "README.md").read_text(encoding="utf-8")
    spec = (pub_dir / "LI-RC-spec.md").read_text(encoding="utf-8")
    assert "NOT YET PUBLISHED" not in readme
    assert "NOT YET PUBLISHED" not in spec
    assert "READY TO PUBLISH" in readme and "approved by RJL 2026-07-07" in readme
    assert "READY TO PUBLISH" in spec and "approved by RJL 2026-07-07" in spec
    assert "SHA-256" in spec
    assert "internal_variant" in readme  # the exclusion is documented, not silent
    # human-push instruction survives the rewrite -- this is a build, not a publish
    assert "human must still" in readme.lower()
    assert "human must still" in spec.lower()


def test_build_public_spec_is_rerunnable_idempotent(built_public_spec_dir):
    pub_dir, _stdout = built_public_spec_dir
    repo_copy = pub_dir.parent.parent
    before = {p.name: p.read_bytes() for p in pub_dir.iterdir()}
    result = subprocess.run(
        [sys.executable, str(repo_copy / "scripts" / "build_public_spec.py")],
        capture_output=True, text=True, cwd=str(repo_copy))
    assert result.returncode == 0
    after = {p.name: p.read_bytes() for p in pub_dir.iterdir()}
    assert before == after, "re-running the build script must be a no-op on unchanged inputs"


def test_reference_scorer_reproduces_expected_score_exactly(built_public_spec_dir):
    pub_dir, _stdout = built_public_spec_dir
    expected = (pub_dir / "sample_results_expected_output.txt").read_text(encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(pub_dir / "reference_scorer.py"),
         str(pub_dir / "sample_results.csv"), str(pub_dir / "scorecard.yaml"),
         "duration_estimating"],
        capture_output=True, text=True, cwd=str(pub_dir))
    assert result.returncode == 0
    assert result.stdout == expected
    # and the hand-computed number itself, so a drift in the sample data or
    # the curve formula fails loudly here rather than only in a byte diff
    assert "84.80 / 100" in result.stdout


def test_build_public_spec_strips_internal_variant_when_source_has_provocative_members(
        built_public_spec_dir):
    """Guard against the parallel N16-N20 wave landing content in
    internal_variant and the strip silently failing to notice: the source
    file must actually HAVE the member entries (fixture assumption), and
    they must be completely absent -- not just weight-zeroed -- from the
    published copy.  Matched as YAML keys (``N16:`` etc.), not a bare
    substring, since the file's own top-of-file explanatory comment
    legitimately names the indices in prose without leaking any privileged
    member data."""
    pub_dir, _stdout = built_public_spec_dir
    src_path = os.path.join(ROOT, "src", "scheduleiq", "scorecard.yaml")
    with open(src_path, encoding="utf-8") as f:
        src_text = f.read()
    for member in ("N16:", "N17:", "N18:", "N19:", "N20:"):
        assert member in src_text  # fixture assumption
    pub_text = (pub_dir / "scorecard.yaml").read_text(encoding="utf-8")
    for member in ("N16:", "N17:", "N18:", "N19:", "N20:"):
        assert member not in pub_text
