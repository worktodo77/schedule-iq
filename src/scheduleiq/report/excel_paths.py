"""Excel workbook for the multi-path analytics (backlog A1, P1-P4).

One workbook, four sheets — the driving-path fingerprint for the latest update,
the top-N float paths, the true merge-point ranking, and the per-update-pair
path-stability register.  Styling reuses the LI helpers from report/excel.py
(teal #1F6F7B headers, gray grid) so these sheets sit alongside the results and
trend workbooks without a visual seam.  excel.py itself is never modified.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from ..analytics.paths import run_path_analytics
from .excel import STATUS_FILL, TEAL, _header, _row


def _fmt(x, nd=1):
    return "" if x is None else round(x, nd)


def _dp_rows(dp):
    rows = []
    for st in dp.steps:
        a = st.activity
        rel = (f"{st.driving_rel.rtype.value} "
               f"lag {st.lag_hours:g}h") if st.driving_rel else "— (target)"
        rows.append([a.code, a.name, rel,
                     "yes" if st.date_satisfied else "float-picked",
                     st.calendar_name, st.constraint or "—",
                     _fmt(st.total_float_days), f"{st.pct_complete:.0f}%"])
    return rows


def write_paths_workbook(series_analysis, paths_results, path: str) -> str:
    """Write the path-analysis workbook.  ``paths_results`` may be a prebuilt
    bundle from run_path_analytics(); if None it is computed here."""
    pr = paths_results or run_path_analytics(series_analysis)
    wb = Workbook()

    # ---- Driving Path ---------------------------------------------------
    ws = wb.active
    ws.title = "Driving Path"
    dp = pr["driving"]
    tgt = dp.target.code if dp.target else "—"
    ws["A1"] = f"Driving path to {tgt} — latest update"
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    meta = [("Target", tgt), ("Tolerance (h)", dp.tolerance_hours),
            ("Path length (activities)", len(dp.steps)),
            ("Longest-path flag agreement",
             "—" if dp.flag_agreement_pct is None else f"{dp.flag_agreement_pct:.0f}%"),
            ("Flag disagreements",
             ", ".join(dp.flag_disagreements) or "—")]
    if dp.reason:
        meta.append(("Note", dp.reason))
    for i, (k, v) in enumerate(meta, 2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = Font(size=10)
    hrow = len(meta) + 3
    _header(ws, hrow, ["Activity", "Name", "Drives next (rel/lag)", "Satisfied",
                       "Calendar", "Constraint", "Total float (d)", "% complete"],
            [12, 34, 20, 14, 20, 18, 14, 12])
    for i, row in enumerate(_dp_rows(dp), hrow + 1):
        _row(ws, i, row)

    # ---- Float Paths ----------------------------------------------------
    ws = wb.create_sheet("Float Paths")
    _header(ws, 1, ["Rank", "Rel. float (d)", "Min float (d)", "Max float (d)",
                    "Activities", "Calendars", "Constraints", "% complete",
                    "Path (codes)"],
            [6, 12, 12, 12, 10, 24, 24, 10, 70])
    for i, fp in enumerate(pr["float_paths"], 2):
        _row(ws, i, [fp.rank, _fmt(fp.rel_float_days), _fmt(fp.min_float_days),
                     _fmt(fp.max_float_days), len(fp.steps),
                     ", ".join(fp.calendars) or "—",
                     ", ".join(fp.constraints) or "—",
                     f"{fp.pct_complete:.0f}%", " → ".join(fp.codes)],
             fill="C6E0B4" if fp.rank == 1 else None)

    # ---- Merge Points ---------------------------------------------------
    ws = wb.create_sheet("Merge Points")
    _header(ws, 1, ["Merge activity", "Name", "Converging near-crit chains",
                    "Tightness — min float (d)", "Converging predecessors"],
            [14, 34, 24, 22, 44])
    for i, m in enumerate(pr["merges"], 2):
        _row(ws, i, [m.code, m.activity.name, m.converging_chains,
                     _fmt(m.tightness_days), ", ".join(m.predecessor_codes)])

    # ---- Path Stability -------------------------------------------------
    ws = wb.create_sheet("Path Stability")
    _header(ws, 1, ["Update pair", "Jaccard", "Joined path", "Left path",
                    "Progress-driven", "Revision-driven", "Attribution"],
            [40, 10, 26, 26, 20, 20, 70])
    for i, p in enumerate(pr["stability"], 2):
        _row(ws, i, [f"{p.earlier_label} → {p.later_label}",
                     "" if p.jaccard is None else round(p.jaccard, 2),
                     ", ".join(p.joined) or "—", ", ".join(p.left) or "—",
                     ", ".join(p.progress_driven) or "—",
                     ", ".join(p.revision_driven) or "—",
                     "  ".join(p.causes)],
             fill=STATUS_FILL.get("WARNING") if (p.joined or p.left) else None)

    wb.save(path)
    return path
