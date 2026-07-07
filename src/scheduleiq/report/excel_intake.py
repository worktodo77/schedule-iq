"""Excel workbook for the intake-accelerator pack (backlog D1-D8).

One LI-styled sheet per accelerator, styling reused from the LI helpers in
report/excel.py (teal #1F6F7B headers, gray grid) so these sheets sit
alongside the results, trend, and path-analysis workbooks without a visual
seam.  excel.py itself is never modified.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from .excel import STATUS_FILL, TEAL, _header, _row


def _fmt(x, nd=1):
    return "" if x is None else round(x, nd)


def _title(ws, text):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)


# --------------------------------------------------------------------------
# D1 — scorecard
# --------------------------------------------------------------------------
def _write_scorecard(wb, sc):
    ws = wb.active
    ws.title = "Scorecard"
    _title(ws, "Data-Completeness Scorecard")
    meta = [
        ("Files", sc.n_files),
        ("Date range", f"{sc.date_range[0]} to {sc.date_range[1]}"
         if sc.date_range[0] else "—"),
        ("Dominant cadence", sc.cadence_label or "—"),
        ("Baseline file", sc.baseline_file or "—"),
        ("Baseline already progressed?",
         "Yes" if sc.baseline_has_progress else "No"),
        ("Format mix", ", ".join(f"{k}: {v}" for k, v in sc.format_mix.items()) or "—"),
        ("Missing months", ", ".join(sc.missing_months) or "—"),
    ]
    if sc.reason:
        meta.append(("Note", sc.reason))
    for i, (k, v) in enumerate(meta, 2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = Font(size=10)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    hrow = len(meta) + 4
    _header(ws, hrow, ["File", "Format", "SCHEDOPTIONS", "% Resourced", "% Cost-Loaded"],
            [30, 12, 14, 14, 14])
    r = hrow + 1
    for f in sc.files:
        _row(ws, r, [f.file, f.format, "yes" if f.has_schedoptions else "no",
                     _fmt(f.pct_resourced, 0), _fmt(f.pct_cost_loaded, 0)],
             fill=None if f.has_schedoptions else STATUS_FILL.get("WARNING"))
        r += 1

    rrow = r + 2
    _header(ws, rrow, ["Topic", "Draft RFI Line"], [18, 100])
    for i, item in enumerate(sc.rfi_items, rrow + 1):
        _row(ws, i, [item.topic, item.text])


# --------------------------------------------------------------------------
# D2 — variance register
# --------------------------------------------------------------------------
def _write_variance(wb, vr):
    ws = wb.create_sheet("Variance")
    _title(ws, f"As-Planned vs As-Built Variance — {vr.baseline_label} vs {vr.current_label}")
    r = 2
    if vr.reason:
        ws.cell(row=r, column=1, value=vr.reason).font = Font(italic=True, size=10)
        r += 2
    _header(ws, r, ["Activity", "Name", "Baseline Start", "Baseline Finish",
                    "Current Start", "Current Finish", "Start Var (wd)",
                    "Finish Var (wd)", "Duration Growth (d)", "On Driving Path"],
            [12, 32, 16, 16, 16, 16, 14, 14, 16, 14])
    r += 1
    for row in vr.rows:
        _row(ws, r, [row.code, row.name, row.baseline_start, row.baseline_finish,
                     row.current_start, row.current_finish, _fmt(row.start_variance_days),
                     _fmt(row.finish_variance_days), _fmt(row.duration_growth_days),
                     "yes" if row.on_driving_path else ""],
             fill="C6E0B4" if row.on_driving_path else None)
        r += 1


# --------------------------------------------------------------------------
# D3 — float ledger
# --------------------------------------------------------------------------
def _write_float_ledger(wb, fl):
    ws = wb.create_sheet("Float Ledger")
    _title(ws, "Float Consumption Ledger")
    r = 2
    if fl.reason:
        ws.cell(row=r, column=1, value=fl.reason).font = Font(italic=True, size=10)
        r += 2

    ws.cell(row=r, column=1, value="Per-activity deltas").font = Font(bold=True, size=11)
    r += 1
    _header(ws, r, ["Update Pair", "Activity", "Name", "WBS", "TF Delta (d)", "Band"],
            [30, 12, 30, 14, 12, 16])
    r += 1
    for row in fl.rows:
        _row(ws, r, [row.pair_label, row.code, row.name, row.wbs,
                     _fmt(row.delta_days), row.band])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="By WBS node").font = Font(bold=True, size=11)
    r += 1
    _header(ws, r, ["Update Pair", "WBS", "N", "Total Delta (d)", "Mean Delta (d)"],
            [30, 14, 8, 16, 16])
    r += 1
    for a in fl.by_wbs:
        _row(ws, r, [a.pair_label, a.key, a.n_activities, _fmt(a.total_delta_days),
                     _fmt(a.mean_delta_days)])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="By float band").font = Font(bold=True, size=11)
    r += 1
    _header(ws, r, ["Update Pair", "Band", "N", "Total Delta (d)", "Mean Delta (d)"],
            [30, 16, 8, 16, 16])
    r += 1
    for a in fl.by_band:
        _row(ws, r, [a.pair_label, a.key, a.n_activities, _fmt(a.total_delta_days),
                     _fmt(a.mean_delta_days)])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Erosion by window").font = Font(bold=True, size=11)
    r += 1
    _header(ws, r, ["Update", "Min Float (d)", "Mean Float (d)", "Consumed (d)"],
            [30, 14, 14, 14])
    r += 1
    for w in fl.erosion_by_window:
        _row(ws, r, [w.label, _fmt(w.min_float_days), _fmt(w.mean_float_days),
                     _fmt(w.consumed_days)])
        r += 1


# --------------------------------------------------------------------------
# D4 — windows
# --------------------------------------------------------------------------
def _write_windows(wb, wp):
    ws = wb.create_sheet("Windows")
    _title(ws, "Proposed Analysis Windows")
    r = 2
    if wp.reason:
        ws.cell(row=r, column=1, value=wp.reason).font = Font(italic=True, size=10)
        r += 2
    _header(ws, r, ["Start DD", "End DD", "Updates in Window", "Driving-Path Summary",
                    "Boundary Kept — Reason"], [14, 14, 34, 50, 60])
    for i, b in enumerate(wp.boundaries, r + 1):
        _row(ws, i, [b.start_dd, b.end_dd, ", ".join(b.labels),
                     b.driving_path_summary, b.kept_reason])


# --------------------------------------------------------------------------
# D5 — concurrency
# --------------------------------------------------------------------------
def _write_concurrency(wb, cc):
    ws = wb.create_sheet("Concurrency")
    _title(ws, "Concurrency Screening — " + cc.caption)
    r = 2
    if cc.reason:
        ws.cell(row=r, column=1, value=cc.reason).font = Font(italic=True, size=10)
        r += 2
    _header(ws, r, ["Window", "Path A", "A Float Δ (d)", "A Finish Slip (d)",
                    "Path B", "B Float Δ (d)", "B Finish Slip (d)"],
            [30, 34, 14, 16, 34, 14, 16])
    for i, c in enumerate(cc.candidates, r + 1):
        _row(ws, i, [c.window_label, " → ".join(c.path_a_codes),
                     _fmt(c.path_a_float_delta_days), _fmt(c.path_a_finish_slip_days),
                     " → ".join(c.path_b_codes), _fmt(c.path_b_float_delta_days),
                     _fmt(c.path_b_finish_slip_days)])


# --------------------------------------------------------------------------
# D6 — events
# --------------------------------------------------------------------------
def _write_events(wb, ev):
    ws = wb.create_sheet("Events")
    _title(ws, "Delay-Event Mapper")
    r = 2
    if ev.reason:
        ws.cell(row=r, column=1, value=ev.reason).font = Font(italic=True, size=10)
        r += 2
    _header(ws, r, ["Event", "Title", "Start", "Finish", "Keywords", "Responsibility",
                    "Schedule Used", "Candidate Activity", "Match Reasons"],
            [10, 26, 12, 12, 20, 14, 20, 16, 40])
    r += 1
    for e in ev.events:
        if not e.matches:
            _row(ws, r, [e.event_id, e.title, e.start, e.finish, "; ".join(e.keywords),
                         e.responsibility, e.schedule_label, "— (no candidates)", e.reason])
            r += 1
            continue
        for m in e.matches:
            _row(ws, r, [e.event_id, e.title, e.start, e.finish, "; ".join(e.keywords),
                         e.responsibility, e.schedule_label, m.activity_code,
                         "; ".join(m.reasons)])
            r += 1


# --------------------------------------------------------------------------
# D7 — responsibility
# --------------------------------------------------------------------------
def _write_responsibility(wb, rr):
    ws = wb.create_sheet("Responsibility")
    _title(ws, "Responsibility Overlay — " + rr.caption)
    r = 2
    if rr.reason:
        ws.cell(row=r, column=1, value=rr.reason).font = Font(italic=True, size=10)
        r += 2
    ws.cell(row=r, column=1, value="Tagged activities (latest update)"
            ).font = Font(bold=True, size=11)
    r += 1
    _header(ws, r, ["Activity", "Name", "WBS", "Party"], [12, 32, 16, 14])
    r += 1
    for a in rr.by_activity:
        _row(ws, r, [a.code, a.name, a.wbs, a.party])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Float erosion by party, per window"
            ).font = Font(bold=True, size=11)
    r += 1
    _header(ws, r, ["Update Pair", "Party", "N", "Total Δ (d)", "Mean Δ (d)"],
            [30, 14, 8, 14, 14])
    r += 1
    for a in rr.aggregates:
        _row(ws, r, [a.window_label, a.party, a.n_activities,
                     _fmt(a.total_float_delta_days), _fmt(a.mean_float_delta_days)])
        r += 1


# --------------------------------------------------------------------------
# D8 — evergreen
# --------------------------------------------------------------------------
def _write_evergreen(wb, eg):
    ws = wb.create_sheet("Evergreen")
    _title(ws, "Evergreen-Activity Detector")
    r = 2
    if eg.reason:
        ws.cell(row=r, column=1, value=eg.reason).font = Font(italic=True, size=10)
        r += 2
    _header(ws, r, ["Activity", "Name", "% Increase", "RD Change (h)",
                    "Finish Moved Earlier?", "History"],
            [12, 30, 12, 14, 18, 70])
    for i, a in enumerate(eg.activities, r + 1):
        hist = "; ".join(f"{h.label}: {h.pct_complete:.0f}% / RD {h.remaining_duration_hours:.0f}h"
                         for h in a.history)
        _row(ws, i, [a.code, a.name, _fmt(a.pct_increase_total, 0),
                     _fmt(a.remaining_duration_change_hours),
                     "yes" if a.forecast_finish_moved_earlier else "no", hist])


def write_intake_workbook(sa, intake_results, path: str) -> str:
    """Write the intake-accelerator workbook: one sheet per D1-D8 output."""
    wb = Workbook()
    _write_scorecard(wb, intake_results.scorecard)
    _write_variance(wb, intake_results.variance)
    _write_float_ledger(wb, intake_results.float_ledger)
    _write_windows(wb, intake_results.windows)
    _write_concurrency(wb, intake_results.concurrency)
    _write_events(wb, intake_results.events)
    _write_responsibility(wb, intake_results.responsibility)
    _write_evergreen(wb, intake_results.evergreen)
    wb.save(path)
    return path
