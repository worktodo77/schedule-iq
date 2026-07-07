"""Tests for the forensic outputs wave — half-step (D9), daily ledger (N3),
robustness certificate (N4), and schedule risk analysis / Monte Carlo (M3).

Exercises the two workbook writers (report/excel_forensic.py,
report/excel_sra.py), the five figures (report/forensic_figures.py,
report/sra_figures.py), and the additive runner.py wiring: the
demo_hs1.xer/demo_hs2.xer pair (both handshake at 100%; see
tests/fixtures/make_fixtures.py's "MIP 3.4 half-step fixture pair" header for
the designed decomposition: progress +9 wd, revision +18 wd, total +27 wd)
produces the full new artifact set with no refusal message, while the legacy
demo_baseline/demo_update1/demo_update2 series (which fails engine
validation, per test_impact_outputs.py) completes with SET-02 refusal
messages and NO new forensic/SRA artifacts — every pre-existing artifact for
that series must still be produced (test_impact_outputs.py already covers
that pre-existing set; this module only asserts the new artifacts are
absent).
"""
import os
import subprocess
import sys

import openpyxl
import pytest

SRC = os.path.join(os.path.dirname(__file__), "..", "src")
sys.path.insert(0, os.path.abspath(SRC))

from scheduleiq.ingest import load                                          # noqa: E402
from scheduleiq.analytics.halfstep import run_halfstep_series               # noqa: E402
from scheduleiq.analytics.dailyledger import run_daily_ledger               # noqa: E402
from scheduleiq.analytics.robustness import run_robustness_certificate      # noqa: E402
from scheduleiq.analytics.montecarlo import (TemplateRule, UncertaintySpec, # noqa: E402
                                             run_simulation)
from scheduleiq.cpm.handshake import clear_handshake_cache                  # noqa: E402
from scheduleiq.report.excel_forensic import write_forensic_workbook        # noqa: E402
from scheduleiq.report.excel_sra import write_sra_workbook                  # noqa: E402
from scheduleiq.report.forensic_figures import (dailyledger_figure,         # noqa: E402
                                                 halfstep_figure,
                                                 robustness_figure)
from scheduleiq.report.sra_figures import scurve_figure, tornado_figure     # noqa: E402
import scheduleiq.runner as runner_mod                                      # noqa: E402
from scheduleiq.runner import run                                           # noqa: E402

FIX = os.path.join(os.path.dirname(__file__), "fixtures")
HS1 = os.path.join(FIX, "demo_hs1.xer")
HS2 = os.path.join(FIX, "demo_hs2.xer")
BASELINE = os.path.join(FIX, "demo_baseline.xer")
U1 = os.path.join(FIX, "demo_update1.xer")
U2 = os.path.join(FIX, "demo_update2.xer")

FORENSIC_SHEETS = {"Half-Step (MIP 3.4)", "Revision Attribution", "MIP 3.3 As-Is",
                   "Daily Ledger", "Robustness Certificate", "Disclosures"}
SRA_SHEETS = {"Summary", "Criticality & Cruciality", "Input Provenance",
             "Risk Events", "Sample", "Disclosures"}

_SRA_TEST_ITERATIONS = 200


@pytest.fixture(scope="session", autouse=True)
def fixtures():
    if not (os.path.exists(HS1) and os.path.exists(HS2)):
        subprocess.run([sys.executable, os.path.join(FIX, "make_fixtures.py")],
                       check=True)


@pytest.fixture(autouse=True)
def _fresh_handshake_cache():
    clear_handshake_cache()
    yield
    clear_handshake_cache()


@pytest.fixture(scope="session")
def hs_schedules():
    return load(HS1)[0], load(HS2)[0]


@pytest.fixture(scope="session")
def hs_dicts(hs_schedules):
    hs1, hs2 = hs_schedules
    return [r.to_dict() for r in run_halfstep_series([hs1, hs2], handshake="require")]


@pytest.fixture(scope="session")
def ledger_dict(hs_schedules):
    hs1, hs2 = hs_schedules
    return run_daily_ledger(hs1, hs2, handshake="require").to_dict()


@pytest.fixture(scope="session")
def cert_dict(hs_schedules):
    hs1, hs2 = hs_schedules
    return run_robustness_certificate([hs1, hs2], handshake="require").to_dict()


@pytest.fixture(scope="session")
def sim_dict(hs_schedules):
    _hs1, hs2 = hs_schedules
    spec = UncertaintySpec(templates=[TemplateRule(match="", low_pct=-10.0, high_pct=10.0)])
    sim = run_simulation(hs2, spec=spec, iterations=_SRA_TEST_ITERATIONS, seed=42,
                         handshake="require")
    return sim.to_dict()


def _all_cell_values(ws) -> list:
    vals = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                vals.append(cell.value)
    return vals


def _wb_all_values(path) -> dict:
    wb = openpyxl.load_workbook(path)
    return {name: _all_cell_values(wb[name]) for name in wb.sheetnames}


# --------------------------------------------------------------- forensic workbook
def test_forensic_workbook_has_expected_sheets(hs_dicts, ledger_dict, cert_dict, tmp_path):
    path = write_forensic_workbook(hs_dicts, [ledger_dict], cert_dict,
                                   str(tmp_path / "forensic.xlsx"))
    assert os.path.exists(path)
    assert os.path.getsize(path) > 1000
    wb = openpyxl.load_workbook(path)
    assert FORENSIC_SHEETS <= set(wb.sheetnames)


def test_forensic_workbook_preliminary_stamp_every_sheet(hs_dicts, ledger_dict,
                                                          cert_dict, tmp_path):
    path = write_forensic_workbook(hs_dicts, [ledger_dict], cert_dict,
                                   str(tmp_path / "forensic.xlsx"))
    wb = openpyxl.load_workbook(path)
    for name in wb.sheetnames:
        ws = wb[name]
        found = any(cell.value and "PRELIMINARY" in str(cell.value)
                   for row in ws.iter_rows(max_row=4) for cell in row)
        assert found, f"sheet {name!r} missing a PRELIMINARY stamp in its header rows"


def test_forensic_workbook_decomposition_identity_row(hs_dicts, ledger_dict, cert_dict,
                                                       tmp_path):
    """The designed half-step fixture: progress +9, revision +18, total +27 wd
    (tests/fixtures/make_fixtures.py header; tests/test_halfstep.py docstring)."""
    path = write_forensic_workbook(hs_dicts, [ledger_dict], cert_dict,
                                   str(tmp_path / "forensic.xlsx"))
    values = _wb_all_values(path)
    hs_vals = values["Half-Step (MIP 3.4)"]
    for expected in (9, 18, 27):
        assert expected in hs_vals, f"{expected} wd not found in Half-Step sheet"


def test_forensic_workbook_ledger_sum_equals_endpoint(hs_dicts, ledger_dict, cert_dict,
                                                       tmp_path):
    path = write_forensic_workbook(hs_dicts, [ledger_dict], cert_dict,
                                   str(tmp_path / "forensic.xlsx"))
    values = _wb_all_values(path)
    ac = ledger_dict["arithmetic_check"]
    assert ac["exact"] is True
    assert ac["sum_of_daily_deltas_wd"] == ac["endpoint_delta_wd"]
    ledger_vals = [str(v) for v in values["Daily Ledger"]]
    assert str(ac["sum_of_daily_deltas_wd"]) in ledger_vals
    assert any("exact: True" in v for v in ledger_vals)


def test_forensic_workbook_certificate_six_variants(hs_dicts, ledger_dict, cert_dict,
                                                     tmp_path):
    assert cert_dict["total_variant_count"] == 6
    path = write_forensic_workbook(hs_dicts, [ledger_dict], cert_dict,
                                   str(tmp_path / "forensic.xlsx"))
    values = _wb_all_values(path)
    variant_ids = {v["variant_id"] for v in cert_dict["variants"]}
    rc_vals = set(str(v) for v in values["Robustness Certificate"])
    assert len(variant_ids) == 6
    for vid in variant_ids:
        assert vid in rc_vals


def test_forensic_workbook_no_responsibility_sheet_without_overlay(hs_dicts, ledger_dict,
                                                                    cert_dict, tmp_path):
    path = write_forensic_workbook(hs_dicts, [ledger_dict], cert_dict,
                                   str(tmp_path / "forensic.xlsx"))
    wb = openpyxl.load_workbook(path)
    assert "Responsibility Subtotals" not in wb.sheetnames


def test_forensic_workbook_handles_none_certificate(hs_dicts, ledger_dict, tmp_path):
    path = write_forensic_workbook(hs_dicts, [ledger_dict], None,
                                   str(tmp_path / "forensic_nocert.xlsx"))
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert "Robustness Certificate" in wb.sheetnames


def test_forensic_workbook_deterministic_on_reread(hs_dicts, ledger_dict, cert_dict,
                                                    tmp_path):
    p1 = write_forensic_workbook(hs_dicts, [ledger_dict], cert_dict, str(tmp_path / "a.xlsx"))
    p2 = write_forensic_workbook(hs_dicts, [ledger_dict], cert_dict, str(tmp_path / "b.xlsx"))
    v1, v2 = _wb_all_values(p1), _wb_all_values(p2)
    assert v1.keys() == v2.keys()
    for name in v1:
        assert v1[name] == v2[name], f"sheet {name!r} differs between two writes"


# --------------------------------------------------------------------- SRA workbook
def test_sra_workbook_has_expected_sheets(sim_dict, tmp_path):
    path = write_sra_workbook(sim_dict, str(tmp_path / "sra.xlsx"))
    assert os.path.exists(path)
    assert os.path.getsize(path) > 1000
    wb = openpyxl.load_workbook(path)
    assert SRA_SHEETS <= set(wb.sheetnames)


def test_sra_workbook_percentiles_present(sim_dict, tmp_path):
    path = write_sra_workbook(sim_dict, str(tmp_path / "sra.xlsx"))
    values = _wb_all_values(path)
    summary_vals = [str(v) for v in values["Summary"]]
    for key in ("P10", "P50", "P80", "P90"):
        assert key in summary_vals
    for key in ("P10", "P50", "P80", "P90"):
        blk = sim_dict["percentiles"][key]
        assert blk["date"] is not None
        assert blk["date"] in summary_vals


def test_sra_workbook_deterministic_on_reread(sim_dict, tmp_path):
    p1 = write_sra_workbook(sim_dict, str(tmp_path / "a.xlsx"))
    p2 = write_sra_workbook(sim_dict, str(tmp_path / "b.xlsx"))
    v1, v2 = _wb_all_values(p1), _wb_all_values(p2)
    assert v1.keys() == v2.keys()
    for name in v1:
        assert v1[name] == v2[name], f"sheet {name!r} differs between two writes"


# ------------------------------------------------------------------------- figures
def test_halfstep_figure_nonempty(hs_dicts, tmp_path):
    p = halfstep_figure(hs_dicts[-1], str(tmp_path / "hs.png"))
    assert os.path.exists(p) and os.path.getsize(p) > 1000


def test_dailyledger_figure_nonempty(ledger_dict, tmp_path):
    p = dailyledger_figure(ledger_dict, str(tmp_path / "dl.png"))
    assert os.path.exists(p) and os.path.getsize(p) > 1000


def test_robustness_figure_nonempty(cert_dict, tmp_path):
    p = robustness_figure(cert_dict, str(tmp_path / "rb.png"))
    assert os.path.exists(p) and os.path.getsize(p) > 500


def test_scurve_figure_nonempty(sim_dict, tmp_path):
    p = scurve_figure(sim_dict, str(tmp_path / "sc.png"))
    assert os.path.exists(p) and os.path.getsize(p) > 1000


def test_tornado_figure_nonempty(sim_dict, tmp_path):
    p = tornado_figure(sim_dict, str(tmp_path / "td.png"))
    assert os.path.exists(p) and os.path.getsize(p) > 500


def test_halfstep_figure_degrades_on_refusal(tmp_path):
    b = load(BASELINE)[0]
    u1 = load(U1)[0]
    res = run_halfstep_series([b, u1], handshake="require")
    assert res[0].refused
    p = halfstep_figure(res[0].to_dict(), str(tmp_path / "hs_refused.png"))
    assert os.path.exists(p) and os.path.getsize(p) > 300


def test_figures_deterministic(hs_dicts, ledger_dict, cert_dict, sim_dict, tmp_path):
    pairs = [
        (halfstep_figure, hs_dicts[-1], "hs"),
        (dailyledger_figure, ledger_dict, "dl"),
        (robustness_figure, cert_dict, "rb"),
        (scurve_figure, sim_dict, "sc"),
        (tornado_figure, sim_dict, "td"),
    ]
    for fn, d, name in pairs:
        p1 = fn(d, str(tmp_path / f"{name}1.png"))
        p2 = fn(d, str(tmp_path / f"{name}2.png"))
        with open(p1, "rb") as f1, open(p2, "rb") as f2:
            assert f1.read() == f2.read(), f"{name} figure not byte-identical"


# ----------------------------------------------------------------------- runner e2e
def test_runner_end_to_end_hs_pair(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    out = str(tmp_path / "out_hs")
    rr = run([HS1, HS2], out, make_pdf=False)
    assert not any("skipped" in m or "refused" in m for m in rr.messages), rr.messages

    expected = ["forensic_diagnostics.xlsx", "fig_halfstep.png", "fig_daily_ledger.png",
               "fig_robustness.png", "sra_diagnostics.xlsx", "fig_sra_scurve.png",
               "fig_sra_tornado.png"]
    for name in expected:
        p = os.path.join(out, name)
        assert os.path.exists(p) and os.path.getsize(p) > 300, f"missing/empty {name}"
        assert p in rr.outputs

    wb = openpyxl.load_workbook(os.path.join(out, "forensic_diagnostics.xlsx"))
    assert FORENSIC_SHEETS <= set(wb.sheetnames)
    wb2 = openpyxl.load_workbook(os.path.join(out, "sra_diagnostics.xlsx"))
    assert SRA_SHEETS <= set(wb2.sheetnames)

    assert os.path.exists(os.path.join(out, "schedule_assessment.docx"))


def test_runner_end_to_end_legacy_series_refuses_gracefully(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    out = str(tmp_path / "out_legacy")
    rr = run([BASELINE, U1, U2], out, make_pdf=False)

    assert any("forensic delay diagnostics refused" in m for m in rr.messages), rr.messages
    assert any("schedule risk analysis refused" in m for m in rr.messages), rr.messages

    for name in ("forensic_diagnostics.xlsx", "fig_halfstep.png", "fig_daily_ledger.png",
                "fig_robustness.png", "sra_diagnostics.xlsx", "fig_sra_scurve.png",
                "fig_sra_tornado.png"):
        assert not os.path.exists(os.path.join(out, name)), f"unexpected artifact {name}"

    # every pre-existing artifact for a legacy series run must still be produced
    for name in ("demo_baseline_results.xlsx", "demo_update1_results.xlsx",
                "demo_update2_results.xlsx", "trend_analysis.xlsx",
                "report_card.xlsx", "score_trace.json", "path_analysis.xlsx",
                "intake_review.xlsx", "statistical_analysis.xlsx",
                "schedule_assessment.docx", "audit/audit_log.jsonl"):
        assert os.path.exists(os.path.join(out, name)), f"missing legacy artifact {name}"


def test_runner_single_file_emits_no_forensic_message(tmp_path, monkeypatch):
    """A single-file run has < 2 schedules: the forensic block must no-op
    silently (no message at all — matching the trend-workbook gating)."""
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    out = str(tmp_path / "out_single")
    rr = run([HS2], out, make_pdf=False)
    assert not any("forensic" in m for m in rr.messages), rr.messages
    assert not os.path.exists(os.path.join(out, "forensic_diagnostics.xlsx"))


def test_runner_determinism_second_run_byte_identical_figures(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_mod, "_SRA_ITERATIONS", _SRA_TEST_ITERATIONS)
    out1 = str(tmp_path / "run1")
    out2 = str(tmp_path / "run2")
    run([HS1, HS2], out1, make_pdf=False)
    run([HS1, HS2], out2, make_pdf=False)

    for name in ("fig_halfstep.png", "fig_daily_ledger.png", "fig_robustness.png",
                "fig_sra_scurve.png", "fig_sra_tornado.png"):
        with open(os.path.join(out1, name), "rb") as f1, \
             open(os.path.join(out2, name), "rb") as f2:
            assert f1.read() == f2.read(), f"{name} not byte-identical across runs"

    wb1 = _wb_all_values(os.path.join(out1, "forensic_diagnostics.xlsx"))
    wb2 = _wb_all_values(os.path.join(out2, "forensic_diagnostics.xlsx"))
    assert wb1 == wb2, "forensic workbook not cell-identical across runs"

    swb1 = _wb_all_values(os.path.join(out1, "sra_diagnostics.xlsx"))
    swb2 = _wb_all_values(os.path.join(out2, "sra_diagnostics.xlsx"))
    assert swb1 == swb2, "SRA workbook not cell-identical across runs"
