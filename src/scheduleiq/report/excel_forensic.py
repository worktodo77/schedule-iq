"""Excel workbook for the forensic outputs wave — half-step (D9), daily
ledger (N3), and robustness certificate (N4).

Consumes the plain-dict serializations — ``HalfStepResult.to_dict()``,
``DailyLedger.to_dict()``, and ``RobustnessCertificate.to_dict()`` — and never
imports the analytics modules directly, matching ``report/excel_impact.py``'s
contract.  Styling reuses the LI helpers from ``report/excel.py`` (teal
#1F6F7B headers, gray grid) so this workbook sits alongside the other
workbooks without a visual seam; ``excel.py`` itself is never modified.

Every sheet is stamped PRELIMINARY in its title row — these are engine-
computed diagnostic deltas, not a competing schedule; the tool-of-record
dates remain the schedule (ADR-0007 §4, presentation rule).
"""
from __future__ import annotations

from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from .excel import STATUS_FILL, TEAL, _header, _row

PRELIM = ("PRELIMINARY — engine diagnostic delta.  Tool-of-record dates "
         "remain the schedule; causation, entitlement, concurrency, and "
         "quantum are reserved to the expert (AACE 29R-03; SCL Protocol "
         "2nd ed.).")


def _fmt(x, nd=1):
    if x is None:
        return ""
    if isinstance(x, bool):
        return "yes" if x else "no"
    if isinstance(x, float):
        return round(x, nd)
    return x


def _title(ws, text: str) -> int:
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    ws["A2"] = PRELIM
    ws["A2"].font = Font(italic=True, size=9, color="C00000")
    return 4


def _meta(ws, r, pairs):
    for k, v in pairs:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=v).font = Font(size=10)
        r += 1
    return r


def _subhead(ws, r, text: str) -> int:
    ws.cell(row=r, column=1, value=text).font = Font(bold=True, size=11, color=TEAL)
    return r + 1


def _pair_label(d: dict[str, Any]) -> str:
    pair = d.get("pair") or {}
    return f"{pair.get('earlier', '—')} -> {pair.get('later', '—')}"


# ---------------------------------------------------------------------------
# Half-Step (MIP 3.4)
# ---------------------------------------------------------------------------
def _write_halfstep(wb, halfstep_dicts: list[dict[str, Any]]):
    ws = wb.active
    ws.title = "Half-Step (MIP 3.4)"
    r = _title(ws, "MIP 3.4 Half-Step Decomposition — Per Update Pair")
    r += 1
    _header(ws, r, ["Pair", "Refused", "E_n Target EF", "H Target EF",
                    "E_n1 Target EF", "Progress (wd)", "Revision (wd)",
                    "Total (wd)", "Identity Holds", "Progress (cd)",
                    "Revision (cd)", "Total (cd)", "Record Move (wd)",
                    "Record Move (cd)", "Record − Engine Δ (wd)",
                    "Computable", "Blocking / Refusal"],
            [26, 10, 16, 16, 16, 14, 14, 12, 14, 14, 14, 12, 16, 16, 18, 12, 60])
    rr = r + 1
    for d in halfstep_dicts:
        decomp = d.get("decomposition") or {}
        eng = decomp.get("engine_dates") or {}
        prog = decomp.get("progress_effect_workdays")
        rev = decomp.get("revision_effect_workdays")
        rec_move = decomp.get("record_movement_workdays")
        eng_delta = (None if (rec_move is None or (prog is None or rev is None))
                    else rec_move - (prog + rev))
        _row(ws, rr, [
            _pair_label(d), _fmt(d.get("refused")),
            eng.get("E_n_target_early_finish"), eng.get("half_step_target_early_finish"),
            eng.get("E_n1_target_early_finish"), prog, rev,
            decomp.get("total_movement_workdays"), _fmt(decomp.get("identity_holds")),
            decomp.get("progress_effect_calendar_days"),
            decomp.get("revision_effect_calendar_days"),
            decomp.get("total_movement_calendar_days"), rec_move,
            decomp.get("record_movement_calendar_days"), eng_delta,
            _fmt(decomp.get("computable")),
            d.get("refusal") or decomp.get("blocking") or ""],
            fill=STATUS_FILL.get("N/A") if (d.get("refused")
                                            or not decomp.get("computable", True))
            else None)
        rr += 1
    ws.auto_filter.ref = f"A{r}:Q{max(rr - 1, r)}"


# ---------------------------------------------------------------------------
# Revision Attribution
# ---------------------------------------------------------------------------
def _write_revision_attribution(wb, halfstep_dicts: list[dict[str, Any]]):
    ws = wb.create_sheet("Revision Attribution")
    r = _title(ws, "Named Revision Attribution (per class, on top of the half-step)")
    r += 1
    _header(ws, r, ["Pair", "Class", "N Edits", "Delta (wd)", "Delta (cd)",
                    "Computable", "Target Finish (engine)", "Blocking"],
            [26, 18, 10, 12, 12, 12, 18, 50])
    rr = r + 1
    for d in halfstep_dicts:
        attrib = d.get("revision_attribution") or {}
        for c in attrib.get("per_class") or []:
            _row(ws, rr, [_pair_label(d), c.get("class"), c.get("n_edits"),
                         c.get("delta_workdays"), c.get("delta_calendar_days"),
                         _fmt(c.get("computable")), c.get("target_finish_engine"),
                         c.get("blocking")],
                 fill=None if c.get("computable") else STATUS_FILL.get("N/A"))
            rr += 1
    rr += 1
    rr = _subhead(ws, rr, "Interaction / residual (sum of classes vs. revision effect)")
    rr += 1
    _header(ws, rr, ["Pair", "Revision Effect (wd)", "Attribution Sum (wd)",
                    "Interaction/Residual (wd)"], [26, 20, 20, 24])
    rr += 1
    for d in halfstep_dicts:
        attrib = d.get("revision_attribution") or {}
        _row(ws, rr, [_pair_label(d), attrib.get("revision_effect_workdays"),
                     attrib.get("attribution_sum_workdays"),
                     attrib.get("interaction_residual_workdays")])
        rr += 1
    rr += 2
    rr = _subhead(ws, rr, "Named top movers (two largest classes per pair)")
    rr += 1
    _header(ws, rr, ["Pair", "Class", "Edit", "Delta (wd)", "Delta (cd)",
                    "Computable", "Target Finish (engine)"],
            [26, 18, 40, 12, 12, 12, 18])
    rr += 1
    for d in halfstep_dicts:
        attrib = d.get("revision_attribution") or {}
        for c in attrib.get("per_class") or []:
            for m in c.get("top_movers") or []:
                if m.get("edit") == "__truncated__":
                    continue
                _row(ws, rr, [_pair_label(d), c.get("class"), m.get("edit"),
                             m.get("delta_workdays"), m.get("delta_calendar_days"),
                             _fmt(m.get("computable")), m.get("target_finish_engine")])
                rr += 1


# ---------------------------------------------------------------------------
# MIP 3.3 As-Is
# ---------------------------------------------------------------------------
def _write_mip33(wb, halfstep_dicts: list[dict[str, Any]]):
    ws = wb.create_sheet("MIP 3.3 As-Is")
    r = _title(ws, "MIP 3.3 Observational Row — Per Update Pair (no bifurcation)")
    r += 1
    _header(ws, r, ["Pair", "Target Code", "Record Move (cd)", "Engine Move (cd)",
                    "Critical-Path Jaccard", "Activities Added", "Activities Deleted",
                    "Logic Δ", "Duration Changes", "Retroactive Actuals",
                    "Constraint Changes"],
            [26, 14, 16, 16, 18, 14, 14, 10, 14, 16, 16])
    rr = r + 1
    for d in halfstep_dicts:
        row = d.get("mip33_row") or {}
        cc = row.get("change_counts") or {}
        _row(ws, rr, [
            _pair_label(d), row.get("target_code"),
            row.get("record_finish_movement_calendar_days"),
            row.get("engine_finish_movement_calendar_days"),
            row.get("critical_path_jaccard"),
            cc.get("activities added"), cc.get("activities deleted"),
            (cc.get("relationships added", 0) + cc.get("relationships deleted", 0)
             + cc.get("relationships modified", 0)),
            cc.get("duration changes"), cc.get("retroactive actual-date changes"),
            cc.get("constraint changes")])
        rr += 1


# ---------------------------------------------------------------------------
# Daily Ledger
# ---------------------------------------------------------------------------
def _write_daily_ledger(wb, ledger_dicts: list[dict[str, Any]]):
    ws = wb.create_sheet("Daily Ledger")
    r = _title(ws, "N3 Daily-Resolution Delay Ledger — Per Update Pair")
    r += 1
    if not ledger_dicts:
        ws.cell(row=r, column=1, value="no daily ledger computed this run"
                ).font = Font(italic=True, size=10)
        return
    for d in ledger_dicts:
        r = _subhead(ws, r, f"Pair: {_pair_label(d)}")
        tgt = d.get("target") or {}
        window = d.get("window") or {}
        ac = d.get("arithmetic_check") or {}
        rec = d.get("reconciliation") or {}
        r = _meta(ws, r, [
            ("Target", tgt.get("code")),
            ("Resolved how", tgt.get("resolved_how")),
            ("Window", f"{window.get('start')} .. {window.get('end_effective')}"
                       f" ({window.get('calendar_days_computed')} cd, "
                       f"capped={window.get('capped')})"),
            ("Computable", _fmt(d.get("computable"))),
            ("Blocking", d.get("blocking") or "—"),
            ("Arithmetic check", f"Σ deltas {ac.get('sum_of_daily_deltas_wd')} wd "
                                 f"== endpoint {ac.get('endpoint_delta_wd')} wd "
                                 f"(exact: {ac.get('exact')})"),
            ("Reconciliation — later as-imported vs. interpolated (wd)",
             rec.get("later_as_imported_vs_interp_wd")),
            ("Reconciliation — earlier as-imported vs. interpolated (wd)",
             rec.get("earlier_as_imported_vs_interp_wd")),
            ("Reconciliation — record movement (wd)", rec.get("record_movement_wd")),
        ])
        r += 1
        if not d.get("computable", True) or not d.get("rows"):
            r += 1
            continue
        _header(ws, r, ["Day", "EF Target", "Delta (wd)", "Cumulative (wd)",
                        "Controlling Activity", "Controlling Party",
                        "Newly Started", "Newly Completed", "Events"],
                [14, 14, 12, 14, 20, 16, 30, 30, 20])
        rr = r + 1
        for row in d.get("rows") or []:
            _row(ws, rr, [
                row.get("day"), row.get("ef_target"), row.get("delta_workdays"),
                row.get("cumulative_workdays"), row.get("controlling_code"),
                row.get("controlling_party"),
                ", ".join(row.get("newly_started") or []),
                ", ".join(row.get("newly_completed") or []),
                ", ".join(row.get("event_ids") or [])],
                fill=STATUS_FILL.get("WARNING") if row.get("delta_workdays") else None)
            rr += 1
        r = rr + 2


# ---------------------------------------------------------------------------
# Responsibility Subtotals
# ---------------------------------------------------------------------------
def _has_responsibility(ledger_dicts: list[dict[str, Any]]) -> bool:
    return any((d.get("responsibility_subtotals") or {}).get("by_party")
              for d in ledger_dicts)


def _write_responsibility(wb, ledger_dicts: list[dict[str, Any]]):
    ws = wb.create_sheet("Responsibility Subtotals")
    r = _title(ws, "Observational Responsibility Subtotals (N3 daily ledger)")
    r += 1
    ws.cell(row=r, column=1, value=(
        "OBSERVATIONAL allocation only: each day's delta is attributed to the "
        "responsibility tag of that day's controlling activity.  A screen, "
        "never an apportionment.")).font = Font(italic=True, size=9)
    r += 2
    _header(ws, r, ["Pair", "Party", "Delta (wd)", "Days"], [26, 24, 14, 10])
    rr = r + 1
    for d in ledger_dicts:
        by_party = (d.get("responsibility_subtotals") or {}).get("by_party") or {}
        for party, v in sorted(by_party.items()):
            _row(ws, rr, [_pair_label(d), party, v.get("delta_workdays"), v.get("days")])
            rr += 1


# ---------------------------------------------------------------------------
# Robustness Certificate
# ---------------------------------------------------------------------------
def _write_robustness(wb, cert: Optional[dict[str, Any]]):
    ws = wb.create_sheet("Robustness Certificate")
    r = _title(ws, "N4 Methodology-Robustness Certificate")
    r += 1
    if cert is None:
        ws.cell(row=r, column=1, value="robustness certificate not computed this run"
                ).font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [
        ("Target", cert.get("target")),
        ("Target resolved how", cert.get("target_resolved_how")),
        ("Schedules in series", cert.get("n_schedules")),
        ("Responsibility overlay applied", _fmt(cert.get("overlay"))),
        ("Computable / total variants",
         f"{cert.get('computable_variant_count')} / {cert.get('total_variant_count')}"),
    ])
    r += 1
    thr = cert.get("verdict_thresholds") or {}
    r = _meta(ws, r, [
        ("STABLE threshold", thr.get("STABLE")),
        ("MODERATE threshold", thr.get("MODERATE")),
        ("UNSTABLE", thr.get("UNSTABLE")),
    ])
    r += 1
    r = _subhead(ws, r, "Stability by series")
    r += 1
    _header(ws, r, ["Series", "N Variants", "Min", "Max", "Range", "Median",
                    "Spread %", "Verdict", "Sentence"],
            [18, 10, 10, 10, 10, 10, 10, 12, 70])
    rr = r + 1
    for s in cert.get("stability") or []:
        _row(ws, rr, [s.get("series"), s.get("n_variants"), s.get("min"),
                     s.get("max"), s.get("range"), s.get("median"),
                     s.get("spread_pct"), s.get("verdict"), s.get("sentence")],
             fill={"STABLE": STATUS_FILL.get("PASS"),
                  "MODERATE": STATUS_FILL.get("WARNING"),
                  "UNSTABLE": STATUS_FILL.get("FAIL")}.get(s.get("verdict")))
        rr += 1
    rr += 2
    rr = _subhead(ws, rr, "Variant grid")
    rr += 1
    _header(ws, rr, ["Variant ID", "Framing", "Statusing", "Boundary", "Contested",
                    "Primary", "Computable", "Total (wd)", "Per Party", "Reason"],
            [30, 20, 16, 18, 16, 10, 12, 12, 40, 50])
    rr2 = rr + 1
    for v in cert.get("variants") or []:
        coord = v.get("coordinates") or {}
        pp = v.get("per_party") or {}
        _row(ws, rr2, [
            v.get("variant_id"), coord.get("framing_label"),
            coord.get("statusing_mode_resolved") or coord.get("statusing"),
            coord.get("boundary"), coord.get("contested"),
            _fmt(v.get("is_primary")), _fmt(v.get("computable")),
            v.get("total_workdays"),
            "; ".join(f"{k}: {val}" for k, val in sorted(pp.items())),
            v.get("reason")],
            fill=None if v.get("computable") else STATUS_FILL.get("N/A"))
        rr2 += 1
    ws.auto_filter.ref = f"A{rr}:J{max(rr2 - 1, rr)}"


# ---------------------------------------------------------------------------
# Disclosures
# ---------------------------------------------------------------------------
def _write_disclosures(wb, halfstep_dicts: list[dict[str, Any]],
                       ledger_dicts: list[dict[str, Any]],
                       cert: Optional[dict[str, Any]]):
    ws = wb.create_sheet("Disclosures")
    r = _title(ws, "Disclosures and Deferred Items")

    def _block(title, items):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11)
        r += 1
        if not items:
            ws.cell(row=r, column=1, value="— none —").font = Font(italic=True, size=10)
            r += 1
        for it in items:
            ws.cell(row=r, column=1, value=str(it)).font = Font(size=10)
            r += 1
        r += 1

    for d in halfstep_dicts:
        _block(f"Half-step disclosures — {_pair_label(d)}", d.get("disclosures") or [])
    for d in ledger_dicts:
        _block(f"Daily ledger disclosures — {_pair_label(d)}", d.get("disclosures") or [])
    if cert is not None:
        _block("Robustness certificate disclosures", cert.get("disclosures") or [])
        _block("Robustness certificate — allocation note",
              [cert.get("allocation_note")] if cert.get("allocation_note") else [])
    ws.column_dimensions["A"].width = 110


def write_forensic_workbook(halfstep_dicts: list[dict[str, Any]],
                            ledger_dicts: list[dict[str, Any]],
                            cert_dict: Optional[dict[str, Any]],
                            out_path: str) -> str:
    """Write the forensic-diagnostics workbook: Half-Step (MIP 3.4), Revision
    Attribution, MIP 3.3 As-Is, Daily Ledger, Responsibility Subtotals (when
    any pair carries a responsibility overlay), Robustness Certificate, and
    Disclosures.  Every argument is a plain-dict serialization from the
    corresponding analytics module's ``to_dict()``; ``cert_dict`` may be
    ``None`` when no robustness certificate was computed this run."""
    wb = Workbook()
    _write_halfstep(wb, halfstep_dicts)
    _write_revision_attribution(wb, halfstep_dicts)
    _write_mip33(wb, halfstep_dicts)
    _write_daily_ledger(wb, ledger_dicts)
    if _has_responsibility(ledger_dicts):
        _write_responsibility(wb, ledger_dicts)
    _write_robustness(wb, cert_dict)
    _write_disclosures(wb, halfstep_dicts, ledger_dicts, cert_dict)
    wb.save(out_path)
    return out_path
