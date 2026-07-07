"""Excel outputs (openpyxl).

Per run: one results workbook per schedule file (summary, all metric results,
findings drill-down per tripped check, calendar inventory, parse log), plus —
for a series — a trend workbook (metric-by-update table, native Excel line
charts of key trends, change register sheets per update pair).

LI palette: teal #1F6F7B header rows, white bold header text, gray grid —
matching the LI report table style.
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..metrics.engine import ScheduleAssessment
from ..trend.series import SeriesAnalysis

TEAL = "1F6F7B"
GRAY = "BFBFBF"
STATUS_FILL = {
    "PASS": "C6E0B4", "FAIL": "F4B084", "WARNING": "FFE699",
    "INFO": "DDEBF7", "N/A": "EDEDED", "NOT EVALUATED": "EDEDED",
}
_thin = Side(style="thin", color=GRAY)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header(ws, row, cols, widths=None):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.border = _border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _row(ws, r, values, fill=None):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = Font(size=10)
        cell.border = _border
        cell.alignment = Alignment(vertical="top", wrap_text=(i > 1))
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)


def write_assessment_workbook(assessment: ScheduleAssessment, path: str) -> str:
    s = assessment.schedule
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"ScheduleIQ Assessment — {s.project_id}"
    ws["A1"].font = Font(bold=True, size=14, color=TEAL)
    meta = [
        ("Source file", s.source_file), ("Format", s.source_format),
        ("SHA-256", s.source_sha256[:16] + "…"),
        ("Data date", s.data_date), ("Project start", s.start_date),
        ("Scheduled finish", s.finish_date), ("Must finish by", s.must_finish_by),
        ("Exported by", s.export_user or "—"), ("Export date", s.export_date),
        ("Activities (real / total)",
         f"{len(s.real_activities)} / {len(s.activities)}"),
        ("Relationships", len(s.relationships)),
        ("Health score (0-100)", assessment.health_score),
    ]
    for k, v in assessment.counts.items():
        if v:
            meta.append((f"Checks {k}", v))
    for i, (k, v) in enumerate(meta, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = Font(size=10)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 44

    ws = wb.create_sheet("Metric Results")
    _header(ws, 1, ["ID", "Check", "Category", "Value", "Threshold",
                    "Threshold source", "Status", "Result", "Findings",
                    "References", "Fuse equivalent"],
            [9, 30, 24, 10, 10, 16, 12, 55, 9, 45, 24])
    r = 2
    for res in assessment.results:
        c = res.check
        thr = ("—" if res.threshold_applied is None else
               (f"{'<=' if c.direction == 'max' else '>='} "
                f"{res.threshold_applied:g}{'%' if c.unit == 'percent' else ''}"))
        _row(ws, r, [c.id, c.name, c.category, res.display_value, thr,
                     res.threshold_source, res.status, res.narrative,
                     len(res.findings), "; ".join(c.references),
                     c.fuse_equivalent or "—"],
             fill=STATUS_FILL.get(res.status))
        r += 1
    ws.auto_filter.ref = f"A1:K{r - 1}"

    ws = wb.create_sheet("Findings")
    _header(ws, 1, ["Check", "Status", "Object ID", "Name", "Detail"],
            [10, 10, 22, 42, 60])
    r = 2
    for res in assessment.results:
        for f in res.findings:
            _row(ws, r, [res.check.id, res.status, f.object_id, f.object_name,
                         f.detail], fill=STATUS_FILL.get(res.status))
            r += 1
    ws.auto_filter.ref = f"A1:E{max(r - 1, 1)}"

    ws = wb.create_sheet("Calendars")
    _header(ws, 1, ["Calendar", "Type", "Hours per day", "Workdays per week",
                    "Holiday exceptions", "Default"], [30, 12, 14, 16, 16, 10])
    for i, c in enumerate(s.calendars.values(), 2):
        _row(ws, i, [c.name or c.uid, c.ctype, c.hours_per_day,
                     c.workdays_per_week, len(c.exceptions_nonwork),
                     "Yes" if c.is_default else ""])

    if s.parse_warnings:
        ws = wb.create_sheet("Parse Log")
        _header(ws, 1, ["Warning"], [110])
        for i, w in enumerate(s.parse_warnings, 2):
            _row(ws, i, [w])

    wb.save(path)
    return path


def write_trend_workbook(sa: SeriesAnalysis, path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Metric Trends"

    labels = sa.labels
    _header(ws, 1, ["ID", "Check", "Unit"] + labels,
            [9, 32, 9] + [18] * len(labels))
    key_rows: dict[str, int] = {}
    r = 2
    matrix_checks = [a.check for a in sa.assessments[0].results]
    for cd in matrix_checks:
        vals = sa.metric_trend(cd.id)
        if all(v is None for v in vals):
            continue
        _row(ws, r, [cd.id, cd.name, cd.unit] +
             [round(v, 2) if v is not None else None for v in vals])
        key_rows[cd.id] = r
        r += 1
    hs_row = r
    _row(ws, r, ["SIQ", "Health score", "index"] +
         [a.health_score for a in sa.assessments])
    r += 1

    def add_chart(title, ids, anchor, y_title):
        ch = LineChart()
        ch.title = title
        ch.height, ch.width = 8, 18
        ch.y_axis.title = y_title
        ch.x_axis.title = "Update"
        for cid in ids:
            rr = hs_row if cid == "SIQ" else key_rows.get(cid)
            if rr is None:
                continue
            data = Reference(ws, min_col=4, max_col=3 + len(labels),
                             min_row=rr, max_row=rr)
            ch.add_data(data, titles_from_data=False)
        cats = Reference(ws, min_col=4, max_col=3 + len(labels), min_row=1)
        ch.set_categories(cats)
        ws.add_chart(ch, anchor)

    row0 = r + 2
    add_chart("Schedule health score by update", ["SIQ"], f"A{row0}", "Score")
    add_chart("DCMA logic and constraint failures (%)",
              ["DCMA-01", "DCMA-05", "DCMA-06"], f"L{row0}", "%")
    add_chart("Negative float and critical share (%)",
              ["DCMA-07", "FLT-01"], f"A{row0 + 18}", "%")

    ws = wb.create_sheet("Series Metrics")
    _header(ws, 1, ["ID", "Check", "Value", "Status", "Result"],
            [9, 34, 10, 12, 80])
    for i, res in enumerate(sa.series_results, 2):
        _row(ws, i, [res.check.id, res.check.name, res.display_value,
                     res.status, res.narrative],
             fill=STATUS_FILL.get(res.status))

    ws = wb.create_sheet("Series Findings")
    _header(ws, 1, ["Check", "Object", "Detail"], [10, 34, 90])
    r = 2
    for res in sa.series_results:
        for f in res.findings:
            _row(ws, r, [res.check.id, f.object_id, f.detail])
            r += 1

    for i, cs in enumerate(sa.changesets):
        ws = wb.create_sheet(f"Changes {i + 1}-{i + 2}")
        ws.cell(row=1, column=1,
                value=f"{cs.earlier.label()}  ->  {cs.later.label()}"
                ).font = Font(bold=True, size=11, color=TEAL)
        _header(ws, 2, ["Change type", "Activity/Pair", "Name", "Before",
                        "After", "Flag"], [26, 22, 38, 26, 26, 14])
        r = 3
        for a in cs.added:
            _row(ws, r, ["Activity added", a.code, a.name, "—", "", ""]); r += 1
        for a in cs.deleted:
            _row(ws, r, ["Activity deleted", a.code, a.name, "", "—", ""]); r += 1
        for ch in (cs.duration_changes + cs.actual_date_changes +
                   cs.planned_date_changes + cs.constraint_changes +
                   cs.calendar_changes + cs.status_changes + cs.name_changes):
            _row(ws, r, [ch.field, ch.code, ch.name, ch.before, ch.after,
                         ch.flag],
                 fill="F4B084" if ch.flag == "RETROACTIVE" else None)
            r += 1
        for lc in cs.logic_changes:
            _row(ws, r, [f"logic {lc.kind}", f"{lc.pred_code} -> {lc.succ_code}",
                         "", "", lc.detail, ""])
            r += 1
        ws.auto_filter.ref = f"A2:F{max(r - 1, 2)}"

    if sa.warnings:
        ws = wb.create_sheet("Series Warnings")
        _header(ws, 1, ["Kind", "Message"], [20, 110])
        for i, w in enumerate(sa.warnings, 2):
            _row(ws, i, [w.kind, w.message])

    wb.save(path)
    return path


def write_benchmark_workbook(assessments: list[ScheduleAssessment], path: str) -> str:
    """Cross-project benchmarking: same metrics, projects side by side."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark"
    labels = [a.schedule.label() for a in assessments]
    _header(ws, 1, ["ID", "Check", "Unit"] + labels,
            [9, 32, 9] + [20] * len(labels))
    r = 2
    for res0 in assessments[0].results:
        vals = []
        for a in assessments:
            rr = a.result(res0.check.id)
            vals.append(round(rr.value, 2) if rr and rr.value is not None else None)
        if all(v is None for v in vals):
            continue
        _row(ws, r, [res0.check.id, res0.check.name, res0.check.unit] + vals)
        r += 1
    _row(ws, r, ["SIQ", "Health score", "index"] +
         [a.health_score for a in assessments])
    wb.save(path)
    return path
