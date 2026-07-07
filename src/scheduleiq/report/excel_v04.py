"""Excel workbook(s) for the v0.4 wiring wave.

``write_v04_workbook`` consumes the plain-dict serializations (``to_dict()``)
produced by the v0.4 analytics modules — the push-button TIA workbench
(``analytics.tia``), the weather overlay (``analytics.weather``), as-built
work-pattern reconstruction (``analytics.workpatterns``), editing-session
forensics (``analytics.editsessions``), and the Ribbon/Phase/Compliance Fuse-
parity analyzers (``analytics.ribbon``/``analytics.phase``/
``analytics.compliance``) — plus the benchmark-corpus context lines
(``analytics.corpus``).  It never imports an analytics module for its logic,
only the already-produced dict, matching ``report/excel_forensic.py``'s and
``report/excel_impact.py``'s contract.  House style reuses the LI helpers
from ``report/excel.py`` (teal header, gray grid); every sheet carries a
PRELIMINARY stamp in its title rows.

``write_internal_workbook`` is a SEPARATE, PRIVILEGED workbook for the five
provocative indices (LI-11..LI-15 — SMI/DDI/ARR/PPS/RSA,
``analytics.li_provocative`` via ``analytics.li_wiring.li_provocative_
results``).  It is never combined with the standard v0.4 workbook and every
sheet carries a prominent PRIVILEGED banner row (ANALYTICS_PROPOSAL §11;
CLAUDE.md §2/§4).
"""
from __future__ import annotations

import json
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..analytics.damages import DamagesConfig, STANDING_LABEL, exposure_for_delta
from .excel import STATUS_FILL, TEAL, _header, _row

PRELIM = ("PRELIMINARY — diagnostic / observational output.  Causation, "
         "entitlement, concurrency, and quantum are reserved to the expert "
         "(CLAUDE.md §4; AACE 29R-03; SCL Protocol 2nd ed.).")

PRIVILEGED_BANNER = (
    "PRIVILEGED AND CONFIDENTIAL — INTERNAL WORK PRODUCT.  Triage indicators "
    "only (\"observations consistent with … that warrants explanation\"); "
    "never a finding of intent or manipulation.  NOT part of the standard "
    "report or artifact set — do not disclose outside privileged review "
    "(ANALYTICS_PROPOSAL §11; CLAUDE.md §2/§4).")


# ---------------------------------------------------------------------------
# shared helpers (mirrors report/excel_forensic.py's house style)
# ---------------------------------------------------------------------------
def _fmt(x, nd=1):
    if x is None:
        return ""
    if isinstance(x, bool):
        return "yes" if x else "no"
    if isinstance(x, float):
        return round(x, nd)
    return x


def _title(ws, text: str, prelim: str = PRELIM) -> int:
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    ws["A2"] = prelim
    ws["A2"].font = Font(italic=True, size=9, color="C00000")
    return 4


def _priv_title(ws, text: str) -> int:
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    ws["A2"] = PRIVILEGED_BANNER
    ws["A2"].font = Font(bold=True, italic=True, size=9, color="C00000")
    ws["A2"].fill = PatternFill("solid", fgColor="FFF3CD")
    ws["A3"] = ""
    return 5


def _meta(ws, r, pairs):
    for k, v in pairs:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=v).font = Font(size=10)
        r += 1
    return r


def _subhead(ws, r, text: str) -> int:
    ws.cell(row=r, column=1, value=text).font = Font(bold=True, size=11, color=TEAL)
    return r + 1


def _kv_block(ws, r, d: dict[str, Any]) -> int:
    """Full-decomposition dump: one row per top-level key, JSON-serializing
    any nested list/dict value (truncated defensively for the Excel cell-
    text limit)."""
    for k in sorted(d):
        v = d[k]
        ws.cell(row=r, column=1, value=str(k)).font = Font(bold=True, size=10)
        if isinstance(v, (dict, list)):
            text = json.dumps(v, sort_keys=True, default=str, ensure_ascii=False)[:30000]
        else:
            text = v
        cell = ws.cell(row=r, column=2, value=text)
        cell.font = Font(size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return r


# ===========================================================================
# write_v04_workbook — TIA / weather / work-pattern / edit-session /
# ribbon / phase / compliance
# ===========================================================================
def _write_summary(wb, parts: dict[str, Any]):
    ws = wb.active
    ws.title = "Summary"
    r = _title(ws, "ScheduleIQ v0.4 Analytics Supplement — Summary")
    r += 1
    r = _subhead(ws, r, "Analyses run this pass")
    r += 1
    analyses_run = parts.get("analyses_run") or []
    if not analyses_run:
        ws.cell(row=r, column=1, value="— none —").font = Font(italic=True, size=10)
        r += 1
    for a in analyses_run:
        ws.cell(row=r, column=1, value=f"• {a}").font = Font(size=10)
        r += 1
    r += 1
    r = _subhead(ws, r, "Benchmark corpus context (ANALYTICS_PROPOSAL §6.9)")
    r += 1
    corpus_lines = parts.get("corpus_lines") or []
    if not corpus_lines:
        ws.cell(row=r, column=1,
               value="— no corpus context this run —").font = Font(italic=True, size=10)
        r += 1
    for line in corpus_lines:
        ws.cell(row=r, column=1, value=line).font = Font(size=10)
        r += 1
    ws.column_dimensions["A"].width = 110
    ws.column_dimensions["B"].width = 40


def _write_tia_impacts(wb, tia: dict[str, Any], damages: Optional[DamagesConfig] = None):
    ws = wb.create_sheet("TIA Impacts")
    r = _title(ws, "Push-Button TIA — Additive Fragnet Impact (MIP 3.6/3.7)")
    r += 1
    tgt = tia.get("target") or {}
    r = _meta(ws, r, [
        ("Target", f"{tgt.get('code')} ({tgt.get('name') or '—'})"),
        ("As-built schedule", tia.get("as_built_schedule_label")),
        ("Mode", tia.get("mode")),
    ])
    r += 1
    for upd in tia.get("updates") or []:
        r = _subhead(ws, r, f"Update: {upd.get('schedule_label')}")
        r += 1
        r = _meta(ws, r, [
            ("Data date", upd.get("data_date")),
            ("Target", (upd.get("target") or {}).get("code")),
            ("Baseline engine target finish", upd.get("baseline_engine_target_finish")),
            ("Record target finish", upd.get("record_target_finish")),
            ("Baseline computable", _fmt(upd.get("baseline_computable"))),
            ("Cumulative total (wd)", upd.get("cumulative_total_workdays")),
            ("Identity holds", _fmt(upd.get("identity_holds"))),
        ])
        r += 1
        cols = ["Event", "Title", "Isolated (wd)", "Marginal (wd)",
                "Cumulative (wd)", "Engine Target Finish", "Computable", "Blocking"]
        widths = [14, 30, 14, 14, 14, 18, 12, 40]
        if damages is not None:
            cols.append("Exposure (cumulative, time-cost)")
            widths.append(30)
        _header(ws, r, cols, widths)
        rr = r + 1
        for row in upd.get("rows") or []:
            vals = [row.get("event_id"), row.get("event_title"),
                   row.get("isolated_delta_workdays"), row.get("marginal_delta_workdays"),
                   row.get("cumulative_delta_workdays"), row.get("engine_target_finish"),
                   _fmt(row.get("computable")), row.get("blocking")]
            if damages is not None:
                exp = exposure_for_delta(row.get("cumulative_delta_workdays"), damages,
                                         "workdays (target calendar, cumulative)")
                vals.append(exp.formula_text)
            _row(ws, rr, vals,
                fill=None if row.get("computable", True) else STATUS_FILL.get("N/A"))
            rr += 1
        r = rr + 2
    if not tia.get("updates"):
        ws.cell(row=r, column=1,
               value="no additive fragnet impact computed this run").font = Font(italic=True, size=10)


def _write_collapse(wb, tia: dict[str, Any], damages: Optional[DamagesConfig] = None):
    ws = wb.create_sheet("Collapse")
    r = _title(ws, "Subtractive Collapsed As-Built (MIP 3.8/3.9)")
    r += 1
    collapse = tia.get("collapse") or {}
    if not collapse:
        ws.cell(row=r, column=1,
               value="no collapse variant computed this run").font = Font(italic=True, size=10)
        return
    cols = ["Party", "Mode", "Blocked", "Calibration OK", "But-For Finish",
           "Compensable (wd)", "Compensable (cd)"]
    widths = [16, 12, 10, 14, 18, 16, 16]
    if damages is not None:
        cols.append("Exposure (time-cost)")
        widths.append(30)
    _header(ws, r, cols, widths)
    rr = r + 1
    for party in sorted(collapse):
        for mode_name in ("global", "stepped"):
            res = (collapse[party] or {}).get(mode_name)
            if res is None:
                continue
            vals = [party, mode_name, _fmt(res.get("is_blocked")),
                   _fmt(res.get("calibration_ok")), res.get("collapsed_finish"),
                   res.get("compensable_days"), res.get("compensable_calendar_days")]
            if damages is not None:
                exp = exposure_for_delta(res.get("compensable_days"), damages,
                                         "workdays (collapse but-for)")
                vals.append(exp.formula_text)
            _row(ws, rr, vals, fill=STATUS_FILL.get("N/A") if res.get("is_blocked") else None)
            rr += 1


def _write_weather_exhibit(wb, weather: dict[str, Any]):
    ws = wb.create_sheet("Weather Exhibit")
    r = _title(ws, "Weather & External-Conditions Overlay (§8.1)")
    r += 1
    r = _meta(ws, r, [
        ("Station", weather.get("station")),
        ("Reference year", weather.get("reference_year")),
        ("Norm years", ", ".join(str(y) for y in weather.get("norm_years") or [])),
        ("Primary calendar", weather.get("primary_calendar")),
        ("Sensitive keywords",
         ", ".join(weather.get("weather_sensitive_keywords") or [])),
    ])
    r += 1
    r = _subhead(ws, r, "Calendar realism — shortfall by month (§8.1a)")
    r += 1
    _header(ws, r, ["Period", "Total Weekdays", "Calendar Embedded Downtime",
                    "Weather Norm Lost", "Shortfall", "Flagged"],
            [12, 14, 20, 16, 12, 10])
    rr = r + 1
    for m in weather.get("calendar_realism") or []:
        _row(ws, rr, [m.get("period"), m.get("total_weekdays"),
                     m.get("calendar_embedded_downtime_days"),
                     m.get("weather_norm_lost_days"), m.get("shortfall_days"),
                     _fmt(m.get("shortfall_flagged"))],
            fill=STATUS_FILL.get("WARNING") if m.get("shortfall_flagged") else None)
        rr += 1
    rr += 2
    rr = _subhead(ws, rr, "Weather-delay exhibit — presented both ways (§8.1c)")
    rr += 1
    _header(ws, rr, ["Period", "Norm Lost (d)", "Actual Lost (d)", "Exceedance (d)",
                     "Affected Near-Critical Activities", "Note"],
            [22, 14, 14, 14, 40, 60])
    rr2 = rr + 1
    for row in weather.get("weather_delay_exhibit") or []:
        affected = ", ".join(a.get("code", "") for a in
                             row.get("affected_weather_sensitive_activities") or [])
        _row(ws, rr2, [row.get("period"), row.get("norm_lost_days"),
                      row.get("actual_lost_days"), row.get("exceedance_days"),
                      affected, row.get("observational_note")],
            fill=(STATUS_FILL.get("N/A") if (row.get("exceedance_days") is None)
                 else STATUS_FILL.get("WARNING") if row.get("exceedance_days", 0) > 0
                 else STATUS_FILL.get("PASS")))
        rr2 += 1


def _write_work_patterns(wb, wp: dict[str, Any]):
    ws = wb.create_sheet("Work Patterns")
    r = _title(ws, "As-Built Work-Pattern Reconstruction (§8.2)")
    r += 1
    if wp.get("reason") and not wp.get("de_facto_calendars"):
        ws.cell(row=r, column=1, value=wp["reason"]).font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [(k, v) for k, v in sorted((wp.get("summary") or {}).items())])
    r += 1
    r = _subhead(ws, r, "De facto calendars — assigned vs. observed")
    r += 1
    _header(ws, r, ["Calendar", "Population (events)", "Assigned Working Days",
                    "Observed Working Days", "Divergence", "Note"],
            [24, 16, 20, 20, 12, 60])
    rr = r + 1
    for p in wp.get("de_facto_calendars") or []:
        _row(ws, rr, [p.get("name"), p.get("population_events"),
                     ", ".join(str(x) for x in p.get("assigned_working_days") or []),
                     ", ".join(str(x) for x in p.get("observed_working_days") or []),
                     _fmt(p.get("divergence")), p.get("note")],
            fill=STATUS_FILL.get("WARNING") if p.get("divergence") else None)
        rr += 1
    rr += 2
    rr = _subhead(ws, rr, "Weekend/non-working events by window")
    rr += 1
    _header(ws, rr, ["Window", "Count"], [40, 10])
    rr2 = rr + 1
    for k, v in sorted((wp.get("weekend_by_window") or {}).items()):
        _row(ws, rr2, [k, v])
        rr2 += 1
    rr2 += 2
    rr2 = _subhead(ws, rr2, "Dormant spans (near/on driving path)")
    rr2 += 1
    _header(ws, rr2, ["Activity", "WBS", "Window", "Span Start", "Span End",
                      "Working Days", "Basis"], [16, 16, 30, 14, 14, 12, 40])
    rr3 = rr2 + 1
    for d in wp.get("dormant_spans") or []:
        _row(ws, rr3, [d.get("activity_code"), d.get("wbs_code"), d.get("window_key"),
                      d.get("span_start"), d.get("span_end"), d.get("working_days"),
                      d.get("basis")])
        rr3 += 1


def _write_edit_sessions(wb, es: dict[str, Any]):
    ws = wb.create_sheet("Edit Sessions")
    r = _title(ws, "Editing-Session Forensics (§6.1) — Flags + Timeline")
    r += 1
    if es.get("reason") and not es.get("timeline"):
        ws.cell(row=r, column=1, value=es["reason"]).font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [(k, v) for k, v in sorted((es.get("summary") or {}).items())])
    r += 1
    r = _subhead(ws, r, "Flagged sessions")
    r += 1
    _header(ws, r, ["Update", "User", "Start", "End", "Activities", "WBS Spread",
                    "Driving-Path Share", "Flags"],
            [14, 20, 18, 18, 12, 12, 16, 60])
    rr = r + 1
    for s in es.get("timeline") or []:
        if not s.get("flags"):
            continue
        flags = "; ".join(f"{f.get('code')}: {f.get('detail')}" for f in s.get("flags") or [])
        _row(ws, rr, [s.get("schedule_label"), s.get("user"), s.get("start_time"),
                     s.get("end_time"), s.get("activity_count"), s.get("wbs_spread"),
                     s.get("driving_path_share"), flags],
            fill=STATUS_FILL.get("WARNING"))
        rr += 1
    rr += 2
    rr = _subhead(ws, rr, "Full session timeline")
    rr += 1
    _header(ws, rr, ["Update", "User", "Start", "End", "Activities", "Flag Count"],
            [14, 20, 18, 18, 12, 12])
    rr2 = rr + 1
    for s in es.get("timeline") or []:
        _row(ws, rr2, [s.get("schedule_label"), s.get("user"), s.get("start_time"),
                      s.get("end_time"), s.get("activity_count"), len(s.get("flags") or [])])
        rr2 += 1
    rr2 += 2
    ws.cell(row=rr2, column=1, value="Innocent explanations").font = Font(bold=True, size=11, color=TEAL)
    rr2 += 1
    for exp in es.get("innocent_explanations") or []:
        ws.cell(row=rr2, column=1, value=exp).font = Font(italic=True, size=9)
        rr2 += 1


def _write_ribbon(wb, ra: dict[str, Any]):
    ws = wb.create_sheet("Ribbon")
    r = _title(ws, "Ribbon Analyzer — Group-by-WBS Quality Concentration (Fuse F1)")
    r += 1
    if ra.get("reason"):
        ws.cell(row=r, column=1, value=ra["reason"]).font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [
        ("Group by", ra.get("group_by")), ("Level", ra.get("level")),
        ("Unassigned activities", ra.get("unassigned_activities")),
        ("Clamped activities", ra.get("clamped_activities")),
    ])
    r += 1
    _header(ws, r, ["Group", "Name", "Activities", "Score", "Worst Checks"],
           [16, 24, 12, 10, 60])
    rr = r + 1
    for row in ra.get("rows") or []:
        worst = "; ".join(f"{cid}:{n}" for cid, n in row.get("worst_checks") or [])
        score = row.get("score", 100.0)
        fill = (STATUS_FILL.get("FAIL") if score < 60
               else STATUS_FILL.get("WARNING") if score < 85 else None)
        _row(ws, rr, [row.get("group"), row.get("group_name"),
                     row.get("activity_count"), score, worst], fill=fill)
        rr += 1


def _write_phase(wb, pa: dict[str, Any]):
    ws = wb.create_sheet("Phase")
    r = _title(ws, "Phase Analyzer — Time-Phased Slicing (Fuse F2)")
    r += 1
    if pa.get("reason"):
        ws.cell(row=r, column=1, value=pa["reason"]).font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [("Bucket", pa.get("bucket")),
                     ("Unbucketed activities", pa.get("unbucketed_activities"))])
    r += 1
    _header(ws, r, ["Period", "Active", "Starting", "Finishing", "Density",
                    "Near-Critical Share", "Constraints"],
           [20, 10, 10, 10, 12, 16, 12])
    rr = r + 1
    for row in pa.get("rows") or []:
        _row(ws, rr, [row.get("bucket_label"), row.get("active_count"),
                     row.get("starting_count"), row.get("finishing_count"),
                     row.get("density"), row.get("near_critical_share"),
                     row.get("constraint_count")])
        rr += 1


def _write_compliance(wb, ca: dict[str, Any]):
    ws = wb.create_sheet("Compliance")
    r = _title(ws, "Per-Period Start/Finish Compliance (Fuse F4)")
    r += 1
    if ca.get("reason"):
        ws.cell(row=r, column=1, value=ca["reason"]).font = Font(italic=True, size=10)
        return
    _header(ws, r, ["Window", "Basis", "Start 0d %", "Start 7d %", "Finish 0d %",
                    "Finish 7d %", "Commitment Reliability %"],
           [26, 40, 12, 12, 12, 12, 20])
    rr = r + 1
    for w in ca.get("windows") or []:
        _row(ws, rr, [f"{w.get('earlier_label')} -> {w.get('later_label')}",
                     w.get("basis"), w.get("start_compliance_pct_0d"),
                     w.get("start_compliance_pct_7d"), w.get("finish_compliance_pct_0d"),
                     w.get("finish_compliance_pct_7d"), w.get("commitment_reliability_pct")])
        rr += 1


def _write_v04_disclosures(wb, parts: dict[str, Any]):
    ws = wb.create_sheet("Disclosures")
    r = _title(ws, "Disclosures and Deferred Items — v0.4 Analytics Supplement")

    def _block(title, items):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11)
        r += 1
        if not items:
            ws.cell(row=r, column=1, value="— none —").font = Font(italic=True, size=10)
            r += 1
        for it in items:
            ws.cell(row=r, column=1, value=str(it)).font = Font(size=10)
            r += 1
        r += 1

    tia = parts.get("tia") or {}
    if tia:
        _block("TIA disclosures", tia.get("disclosures") or [])
    weather = parts.get("weather") or {}
    if weather:
        _block("Weather overlay disclosures", weather.get("disclosures") or [])
    wp = parts.get("workpatterns") or {}
    if wp:
        _block("Work-pattern disclosures", wp.get("disclosures") or [])
    es = parts.get("editsessions") or {}
    if es:
        _block("Edit-session disclosures", es.get("disclosures") or [])
    ra = parts.get("ribbon") or {}
    if ra and ra.get("excluded_checks"):
        _block("Ribbon — excluded checks",
              [f"{cid}: {ra['excluded_reason'].get(cid, '')}"
               for cid in ra["excluded_checks"]])
    pa = parts.get("phase") or {}
    if pa and pa.get("excluded_checks"):
        _block("Phase — excluded checks",
              [f"{cid}: {pa['excluded_reason'].get(cid, '')}"
               for cid in pa["excluded_checks"]])
    ws.column_dimensions["A"].width = 110


def write_v04_workbook(parts: dict[str, Any], out_path: str,
                       damages: Optional[DamagesConfig] = None) -> str:
    """Write the v0.4 analytics-supplement workbook.  ``parts`` carries the
    plain-dict serialization of each analytics module's result (``None`` /
    absent when that analysis did not run this pass), plus ``analyses_run``
    (list[str], for the Summary sheet) and ``corpus_lines`` (list[str], the
    benchmark-corpus context lines).  Sheets are present ONLY when their
    part exists: Summary (always), TIA Impacts, Collapse, Weather Exhibit,
    Work Patterns, Edit Sessions, Ribbon, Phase, Compliance, Disclosures
    (always).  ``damages`` (S7) is OPTIONAL and adds a time-cost exposure
    column to the TIA/Collapse deltas only, matching
    ``report/excel_forensic.py``'s convention."""
    wb = Workbook()
    _write_summary(wb, parts)

    tia = parts.get("tia")
    if tia:
        _write_tia_impacts(wb, tia, damages)
        if tia.get("collapse"):
            _write_collapse(wb, tia, damages)

    weather = parts.get("weather")
    if weather:
        _write_weather_exhibit(wb, weather)

    wp = parts.get("workpatterns")
    if wp:
        _write_work_patterns(wb, wp)

    es = parts.get("editsessions")
    if es:
        _write_edit_sessions(wb, es)

    ra = parts.get("ribbon")
    if ra:
        _write_ribbon(wb, ra)

    pa = parts.get("phase")
    if pa:
        _write_phase(wb, pa)

    ca = parts.get("compliance")
    if ca:
        _write_compliance(wb, ca)

    _write_v04_disclosures(wb, parts)
    wb.save(out_path)
    return out_path


# ===========================================================================
# write_internal_workbook — LI-11..LI-15 provocative indices (PRIVILEGED)
# ===========================================================================
_INDEX_TITLES = [
    ("LI-11", "LI-11 SMI", "SMI — Schedule Manipulation Indicator (§11.1)"),
    ("LI-12", "LI-12 DDI", "DDI — Directed Date Index (§11.2)"),
    ("LI-13", "LI-13 ARR", "ARR — Attribution Robustness Ratio (§11.3)"),
    ("LI-14", "LI-14 PPS", "PPS — Pacing Plausibility Score (§11.4)"),
    ("LI-15", "LI-15 RSA", "RSA — Rebuttal Surface Area (§11.5)"),
]


def _write_index_sheet(ws, title: str, mr) -> None:
    ws.title = title
    r = _priv_title(ws, title)
    r += 1
    if mr is None:
        ws.cell(row=r, column=1,
               value="not computed this run").font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [
        ("Check ID", mr.check.id),
        ("Value", _fmt(mr.value)),
    ])
    r += 1
    ws.cell(row=r, column=1, value="Interpretation").font = Font(bold=True, size=10)
    r += 1
    ws.cell(row=r, column=1, value=mr.narrative or "").font = Font(size=10)
    r += 2
    if mr.findings:
        r = _subhead(ws, r, "Named findings (decomposition, first 25)")
        r += 1
        _header(ws, r, ["Object", "Name", "Detail"], [24, 24, 80])
        rr = r + 1
        for f in mr.findings:
            _row(ws, rr, [f.object_id, f.object_name, f.detail])
            rr += 1
        r = rr + 2
    decomp = getattr(mr, "decomposition", None) or {}
    if decomp:
        r = _subhead(ws, r, "Full decomposition (raw)")
        r += 1
        r = _kv_block(ws, r, decomp)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90


def _write_certificate_sheet(wb, cert: Optional[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Robustness Certificate")
    r = _priv_title(ws, "N4 Methodology-Robustness Certificate (context for ARR/RSA)")
    r += 1
    if cert is None:
        ws.cell(row=r, column=1,
               value="robustness certificate not computed this run").font = Font(italic=True, size=10)
        return
    r = _meta(ws, r, [
        ("Target", cert.get("target")),
        ("Overlay applied", _fmt(cert.get("overlay"))),
        ("Computable / total variants",
         f"{cert.get('computable_variant_count')} / {cert.get('total_variant_count')}"),
    ])
    r += 1
    r = _subhead(ws, r, "Stability sentences")
    r += 1
    for s in cert.get("sentences") or []:
        ws.cell(row=r, column=1, value=s).font = Font(size=10)
        r += 1
    r += 1
    r = _subhead(ws, r, "Stability by series")
    r += 1
    _header(ws, r, ["Series", "N Variants", "Min", "Max", "Range", "Median",
                    "Spread %", "Verdict"], [18, 10, 10, 10, 10, 10, 10, 12])
    rr = r + 1
    for s in cert.get("stability") or []:
        _row(ws, rr, [s.get("series"), s.get("n_variants"), s.get("min"),
                     s.get("max"), s.get("range"), s.get("median"),
                     s.get("spread_pct"), s.get("verdict")],
            fill={"STABLE": STATUS_FILL.get("PASS"),
                 "MODERATE": STATUS_FILL.get("WARNING"),
                 "UNSTABLE": STATUS_FILL.get("FAIL")}.get(s.get("verdict")))
        rr += 1
    ws.column_dimensions["A"].width = 90


def write_internal_workbook(prov_results, certificate: Optional[dict[str, Any]],
                            out_path: str) -> str:
    """PRIVILEGED / INTERNAL workbook for the five provocative indices
    (LI-11..LI-15 — SMI/DDI/ARR/PPS/RSA).  ``prov_results`` is the
    ``list[MetricResult]`` returned by
    ``analytics.li_wiring.li_provocative_results``; ``certificate`` is the
    optional N4 :class:`~scheduleiq.analytics.robustness.RobustnessCertificate`
    ``.to_dict()`` feeding ARR/RSA's context (``None`` when not computable
    this run).  Every sheet carries a prominent PRIVILEGED banner row — this
    workbook is never part of the standard report/artifact set and must
    never be shared outside privileged review (ANALYTICS_PROPOSAL §11;
    CLAUDE.md §2/§4)."""
    by_id = {mr.check.id: mr for mr in (prov_results or [])}
    wb = Workbook()
    first = True
    for cid, short_title, long_title in _INDEX_TITLES:
        ws = wb.active if first else wb.create_sheet()
        first = False
        _write_index_sheet(ws, short_title, by_id.get(cid))
        ws["A4"] = long_title
        ws["A4"].font = Font(italic=True, size=9)
    _write_certificate_sheet(wb, certificate)
    wb.save(out_path)
    return out_path
