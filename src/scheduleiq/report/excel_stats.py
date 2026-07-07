"""Excel workbook for the statistical screens and earned-schedule credibility
analytics (backlog S1, S3).  Four sheets — Statistical Screens (Benford/
round-number/percent-step per schedule), Distribution Drift (K-S per update
pair), Progress Physics (implied rates + P90 findings), and Earned Schedule
(ES(t)/SPI(t)/TSPI(t)/IEAC(t) per update).  Styling reuses the LI helpers from
report/excel.py (teal #1F6F7B headers, gray grid) exactly as excel_paths.py
does; excel.py itself is never modified.

NOT wired into runner.py or report_builder.py in this wave (see the module
docstring of analytics/statistical.py's ``run_stats`` and this module's
``write_stats_workbook`` — both are plain functions the lead integrates
alongside the existing path-analysis workbook call in runner.run(), to avoid
touching runner.py/report_builder.py in the same wave as a concurrent agent's
edits there).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from ..analytics.earned_schedule import EarnedScheduleResult
from .excel import STATUS_FILL, TEAL, _header, _row

_ = STATUS_FILL  # re-exported import kept for sheets that may want fills later


def _fmt(x, nd=2):
    return "" if x is None else round(x, nd)


def _write_statistical_screens(wb: Workbook, stats_results: dict) -> None:
    ws = wb.active
    ws.title = "Statistical Screens"
    ws["A1"] = "Statistical manipulation screens (Benford / round-number / percent-step)"
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    benford = stats_results.get("benford", [])
    ws["A2"] = benford[0].caution if benford else \
        "Statistical screens indicate patterns for review, not proof of manipulation."
    ws["A2"].font = Font(italic=True, size=9)

    r = 4
    _header(ws, r, ["Schedule", "N durations", "Chi² first digit",
                    "Chi² last digit", "Round-5d %", "N in-progress",
                    "Pct-step-5 %", "Note"],
            [30, 12, 16, 16, 12, 14, 14, 40])
    r += 1
    for b in benford:
        _row(ws, r, [b.label, b.n_durations, _fmt(b.chi2_first_digit),
                     _fmt(b.chi2_last_digit), _fmt(b.round5_pct, 1),
                     b.n_in_progress_pct, _fmt(b.pct_step5_pct, 1),
                     b.reason or "—"])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="First-digit distribution (observed % vs Benford expected %)"
           ).font = Font(bold=True, size=10)
    r += 1
    _header(ws, r, ["Schedule"] + [f"digit {d}" for d in range(1, 10)],
            [30] + [10] * 9)
    r += 1
    for b in benford:
        _row(ws, r, [b.label] + [_fmt(b.first_digit_pct.get(d), 1) for d in range(1, 10)])
        r += 1
    _row(ws, r, ["Benford expected"] +
        [_fmt(b.first_digit_expected_pct.get(d), 1) if benford else ""
         for d in range(1, 10)])
    r += 2

    ws.cell(row=r, column=1, value="Last-digit distribution (observed % vs uniform expected %)"
           ).font = Font(bold=True, size=10)
    r += 1
    _header(ws, r, ["Schedule"] + [f"digit {d}" for d in range(0, 10)],
            [30] + [10] * 10)
    r += 1
    for b in benford:
        _row(ws, r, [b.label] + [_fmt(b.last_digit_pct.get(d), 1) for d in range(0, 10)])
        r += 1
    _row(ws, r, ["Uniform expected"] +
        [_fmt(b.last_digit_expected_pct.get(d), 1) if benford else ""
         for d in range(0, 10)])


def _write_distribution_drift(wb: Workbook, stats_results: dict) -> None:
    ws = wb.create_sheet("Distribution Drift")
    ws["A1"] = "Duration-distribution drift between updates (manual two-sample K-S)"
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    _header(ws, 3, ["Update pair", "N common", "K-S (common OD)", "N added",
                    "N incumbent", "K-S (added vs incumbent)", "Narrative"],
            [40, 10, 16, 10, 12, 20, 60])
    r = 4
    for d in stats_results.get("drift", []):
        _row(ws, r, [f"{d.earlier_label} → {d.later_label}", d.n_common,
                     _fmt(d.ks_common), d.n_added, d.n_incumbent,
                     _fmt(d.ks_added), d.narrative])
        r += 1


def _write_progress_physics(wb: Workbook, stats_results: dict) -> None:
    ws = wb.create_sheet("Progress Physics")
    ws["A1"] = "Progress physics: implied production rates vs demonstrated P90"
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    physics = stats_results.get("physics")
    ws["A2"] = physics.narrative if physics else ""
    ws["A2"].font = Font(italic=True, size=9)
    ws["A3"] = physics.caution if physics else ""
    ws["A3"].font = Font(italic=True, size=9)

    r = 5
    ws.cell(row=r, column=1, value="Implied production-rate observations"
           ).font = Font(bold=True, size=10)
    r += 1
    _header(ws, r, ["Activity", "Period", "Rate (units/h)"], [14, 40, 16])
    r += 1
    for rp in (physics.rates if physics else []):
        _row(ws, r, [rp.code, rp.period, _fmt(rp.rate_per_hour, 3)])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Remaining work exceeding a demonstrated P90 rate"
           ).font = Font(bold=True, size=10)
    r += 1
    _header(ws, r, ["Activity", "Name", "Period", "Required rate (units/h)",
                    "Own P90", "Project P90", "Detail"],
            [14, 30, 40, 18, 12, 14, 60])
    r += 1
    for f in (physics.findings if physics else []):
        _row(ws, r, [f.code, f.name, f.period, _fmt(f.required_rate_per_hour, 3),
                     _fmt(f.own_p90, 3), _fmt(f.project_p90, 3), f.detail],
             fill=STATUS_FILL.get("WARNING"))
        r += 1


def _write_earned_schedule(wb: Workbook, es_results: EarnedScheduleResult) -> None:
    ws = wb.create_sheet("Earned Schedule")
    ws["A1"] = "Earned-schedule forecast credibility"
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    meta = [("Baseline", es_results.baseline_label),
            ("Basis", es_results.basis or "—"),
            ("Baseline start", es_results.baseline_start),
            ("Baseline finish (PD target)", es_results.baseline_finish),
            ("Planned duration PD (days)", _fmt(es_results.planned_duration_days, 1))]
    if es_results.reason:
        meta.append(("Note", es_results.reason))
    for i, (k, v) in enumerate(meta, 2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=str(v) if v is not None else "—"
               ).font = Font(size=10)

    hrow = len(meta) + 4
    _header(ws, hrow, ["Update", "Data date", "AT (d)", "PV", "EV", "ES (d)",
                       "ES date", "SPI(t)", "TSPI(t)", "IEAC (d)", "IEAC date",
                       "Interpretation"],
            [30, 14, 10, 10, 10, 10, 14, 10, 10, 10, 14, 60])
    for i, p in enumerate(es_results.points, hrow + 1):
        _row(ws, i, [p.label,
                     p.data_date.strftime("%Y-%m-%d") if p.data_date else "",
                     _fmt(p.at_days, 1), _fmt(p.pv_value, 1), _fmt(p.ev_value, 1),
                     _fmt(p.es_days, 1),
                     p.es_date.strftime("%Y-%m-%d") if p.es_date else "",
                     _fmt(p.spi_t, 2), _fmt(p.tspi_t, 2), _fmt(p.ieac_days, 1),
                     p.ieac_date.strftime("%Y-%m-%d") if p.ieac_date else "",
                     p.interpretation],
             fill=STATUS_FILL.get("WARNING")
             if (p.tspi_t is not None and p.tspi_t > 1.10) else None)


def write_stats_workbook(sa, stats_results: dict,
                         es_results: EarnedScheduleResult, path: str) -> str:
    """Write the statistical-screens + earned-schedule workbook.

    ``sa`` is the SeriesAnalysis (kept for signature symmetry with the other
    report writers and future per-schedule cross-references; the sheets below
    are driven entirely by ``stats_results`` (analytics.statistical.run_stats)
    and ``es_results`` (analytics.earned_schedule.earned_schedule)).
    """
    del sa  # not currently needed directly; see docstring
    wb = Workbook()
    _write_statistical_screens(wb, stats_results)
    _write_distribution_drift(wb, stats_results)
    _write_progress_physics(wb, stats_results)
    _write_earned_schedule(wb, es_results)
    wb.save(path)
    return path
