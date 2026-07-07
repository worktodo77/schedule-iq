"""Excel workbook for the Monte Carlo / schedule risk analysis (SRA) outputs
(backlog M3).

Consumes the plain-dict serialization ``SimulationResult.to_dict()``
(:mod:`scheduleiq.analytics.montecarlo`) and never imports the analytics
module directly, matching ``report/excel_impact.py``'s contract.  Styling
reuses the LI helpers from ``report/excel.py`` (teal #1F6F7B headers, gray
grid) so this workbook sits alongside the other workbooks without a visual
seam; ``excel.py`` itself is never modified.

Every sheet is stamped PRELIMINARY in its title row.  When the simulation
carries a ``branding`` string (the M4 SRA-readiness gate stamped
DIAGNOSTIC_ONLY), the Summary sheet shows it as a prominent red banner row.
"""
from __future__ import annotations

from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..analytics.damages import (DamagesConfig, STANDING_LABEL, exposure_for_date,
                                 exposure_for_delta)
from .excel import STATUS_FILL, TEAL, _header, _row

PRELIM = ("PRELIMINARY — diagnostic output of a schedule risk model.  "
         "Tool-of-record dates remain the schedule; causation, entitlement, "
         "concurrency, and quantum are reserved to the expert (AACE 29R-03; "
         "SCL Protocol 2nd ed.).")

_SAMPLE_ROW_CAP = 2000


def _fmt(x, nd=2):
    if x is None:
        return ""
    if isinstance(x, bool):
        return "yes" if x else "no"
    if isinstance(x, float):
        return round(x, nd)
    return x


def _title(ws, text: str) -> int:
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=13, color=TEAL)
    ws["A2"] = PRELIM
    ws["A2"].font = Font(italic=True, size=9, color="C00000")
    return 4


def _meta(ws, r, pairs):
    for k, v in pairs:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=v).font = Font(size=10)
        r += 1
    return r


def _subhead(ws, r, text: str) -> int:
    ws.cell(row=r, column=1, value=text).font = Font(bold=True, size=11, color=TEAL)
    return r + 1


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
def _write_summary(wb, sim: dict[str, Any], damages: Optional[DamagesConfig] = None):
    ws = wb.active
    ws.title = "Summary"
    tgt = sim.get("target") or {}
    r = _title(ws, f"Schedule Risk Analysis — {tgt.get('code') or tgt.get('uid') or '—'}")

    branding = sim.get("branding")
    if branding:
        ws.cell(row=r, column=1, value=branding).font = Font(bold=True, size=11,
                                                              color="FFFFFF")
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="C00000")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 2

    readiness = sim.get("readiness") or {}
    r = _meta(ws, r, [
        ("Target code", tgt.get("code")),
        ("Target name", tgt.get("name")),
        ("Target calendar", tgt.get("calendar")),
        ("Target resolved how", tgt.get("resolved_how")),
        ("Iterations", sim.get("iterations")),
        ("Seed", sim.get("seed")),
        ("Correlation (rho)", sim.get("correlation")),
        ("Handshake mode", sim.get("handshake_mode")),
    ])
    r += 1
    r = _subhead(ws, r, "M4 SRA-readiness gate")
    r += 1
    r = _meta(ws, r, [
        ("Verdict", readiness.get("verdict")),
        ("Leads", readiness.get("leads")),
        ("Hard constraints", readiness.get("hard_constraints")),
        ("Open ends", readiness.get("open_ends")),
        ("Handshake passed", _fmt(readiness.get("handshake_passed"))),
        ("Screens failed", "; ".join(readiness.get("screens_failed") or []) or "— none —"),
    ])
    r += 1
    r = _subhead(ws, r, "Percentiles vs. deterministic and record")
    r += 1
    pct = sim.get("percentiles") or {}
    r = _meta(ws, r, [
        ("Deterministic engine finish", pct.get("deterministic_engine_finish")),
        ("Record finish", pct.get("record_finish")),
    ])
    r += 1
    cols = ["Percentile", "Offset (wd)", "Date", "vs. Deterministic (wd)"]
    widths = [14, 14, 16, 22]
    if damages is not None:
        cols += ["Time-Cost Exposure (vs. deterministic)", "LD Exposure (vs. contractual)"]
        widths += [30, 30]
    _header(ws, r, cols, widths)
    rr = r + 1
    for key in ("P10", "P50", "P80", "P90"):
        blk = pct.get(key) or {}
        vals = [key, blk.get("offset"), blk.get("date"),
                blk.get("workdays_vs_deterministic")]
        if damages is not None:
            tc = exposure_for_delta(blk.get("workdays_vs_deterministic"), damages,
                                    "workdays (target calendar, vs. deterministic)")
            ld = exposure_for_date(blk.get("date"), damages)
            vals += [tc.formula_text, ld.formula_text]
        _row(ws, rr, vals)
        rr += 1
    rr += 2
    rr = _subhead(ws, rr, "Merge bias (target)")
    rr += 1
    mb = sim.get("merge_bias") or {}
    rr = _meta(ws, rr, [
        ("Scope", mb.get("scope")),
        ("Note", mb.get("note")),
        ("Deterministic offset (wd)", mb.get("deterministic_offset")),
        ("P50 offset (wd)", mb.get("p50_offset")),
        ("Merge bias (wd)", mb.get("merge_bias_workdays")),
    ])


# ---------------------------------------------------------------------------
# Criticality & Cruciality
# ---------------------------------------------------------------------------
def _write_criticality(wb, sim: dict[str, Any]):
    ws = wb.create_sheet("Criticality & Cruciality")
    r = _title(ws, "Per-Activity Criticality and Cruciality (sorted by cruciality)")
    r += 1
    r = _subhead(ws, r, "Cruciality — varying activities only")
    r += 1
    _header(ws, r, ["Code", "Tier", "Cruciality", "Spearman", "Criticality Index (%)"],
            [16, 14, 14, 14, 20])
    rr = r + 1
    for row in sim.get("cruciality") or []:
        _row(ws, rr, [row.get("code"), row.get("tier"), row.get("cruciality"),
                     row.get("spearman"), row.get("criticality_index_pct")])
        rr += 1
    rr += 2
    rr = _subhead(ws, rr, "Criticality index — all incomplete activities")
    rr += 1
    _header(ws, rr, ["Code", "Criticality Index (%)"], [16, 20])
    rr2 = rr + 1
    for row in sim.get("criticality_index") or []:
        _row(ws, rr2, [row.get("code"), row.get("criticality_index_pct")],
             fill=STATUS_FILL.get("WARNING") if (row.get("criticality_index_pct") or 0) >= 50
             else None)
        rr2 += 1
    ws.auto_filter.ref = f"A{r}:E{max(rr - 1, r)}"


# ---------------------------------------------------------------------------
# Input Provenance
# ---------------------------------------------------------------------------
def _write_provenance(wb, sim: dict[str, Any]):
    ws = wb.create_sheet("Input Provenance")
    r = _title(ws, "Per-Activity Input Provenance (three-point > template > empirical)")
    r += 1
    _header(ws, r, ["Code", "Tier", "Base Remaining (wd)", "Params"], [16, 14, 18, 70])
    rr = r + 1
    for row in sim.get("input_provenance") or []:
        params = row.get("params") or {}
        notes = "; ".join(f"{k}: {v}" for k, v in params.items())
        _row(ws, rr, [row.get("code"), row.get("tier"),
                     row.get("base_remaining_workdays"), notes])
        rr += 1


# ---------------------------------------------------------------------------
# Risk Events
# ---------------------------------------------------------------------------
def _write_risk_events(wb, sim: dict[str, Any]):
    ws = wb.create_sheet("Risk Events")
    r = _title(ws, "Risk Register (probabilistic impacts)")
    r += 1
    events = sim.get("risk_events") or []
    if not events:
        ws.cell(row=r, column=1, value="no risk events configured this run"
                ).font = Font(italic=True, size=10)
        return
    _header(ws, r, ["ID", "Probability", "Affected Code", "Affected UID",
                    "Distribution", "Optimistic (d)", "Most Likely (d)",
                    "Pessimistic (d)"],
            [16, 12, 16, 16, 14, 14, 14, 14])
    rr = r + 1
    for ev in events:
        dist = ev.get("impact_dist") or {}
        _row(ws, rr, [ev.get("id"), ev.get("probability"), ev.get("affected_code"),
                     ev.get("affected_uid"), dist.get("dist"),
                     dist.get("optimistic_days"), dist.get("most_likely_days"),
                     dist.get("pessimistic_days")])
        rr += 1


# ---------------------------------------------------------------------------
# Sample
# ---------------------------------------------------------------------------
def _write_sample(wb, sim: dict[str, Any]):
    ws = wb.create_sheet("Sample")
    r = _title(ws, "Target Completion Sample")
    r += 1
    sample = sim.get("target_sample") or {}
    offsets = sample.get("offsets") or []
    dates = sample.get("dates") or []
    n = len(offsets)
    capped = n > _SAMPLE_ROW_CAP
    ws.cell(row=r, column=1, value=(
        f"{n} iteration(s) sampled" +
        (f"; showing the first {_SAMPLE_ROW_CAP} rows (capped for workbook size)"
         if capped else "."))).font = Font(italic=True, size=10)
    r += 2
    _header(ws, r, ["Iteration", "Offset (wd)", "Date"], [12, 14, 16])
    rr = r + 1
    for i in range(min(n, _SAMPLE_ROW_CAP)):
        _row(ws, rr, [i + 1, offsets[i], dates[i] if i < len(dates) else None])
        rr += 1


# ---------------------------------------------------------------------------
# Exposure (damages overlay, backlog S7 — only written when damages given)
# ---------------------------------------------------------------------------
def _write_exposure(wb, sim: dict[str, Any], damages: DamagesConfig):
    ws = wb.create_sheet("Exposure")
    r = _title(ws, "Exposure — Damages/LD Overlay on the Schedule Risk Analysis")
    ws.cell(row=r, column=1, value=STANDING_LABEL).font = Font(italic=True, size=9,
                                                               color="C00000")
    r += 2
    r = _meta(ws, r, [
        ("Currency", damages.currency),
        ("Daily basis (analyst rates)", damages.daily_basis),
        ("Time-related cost rate (per day)", damages.time_cost_per_day),
        ("LD rate (per calendar day)", damages.ld_rate_per_day),
        ("Contractual completion", damages.contractual_completion),
        ("LD math enabled", _fmt(damages.ld_enabled)),
    ])
    r += 1

    pct = sim.get("percentiles") or {}
    r = _subhead(ws, r, "Per-percentile exposure (P-date vs. contractual completion; "
                       "offset vs. deterministic)")
    r += 1
    _header(ws, r, ["Percentile", "Date", "LD formula", "Offset (wd)", "Time-cost formula"],
            [14, 16, 55, 14, 55])
    rr = r + 1
    for key in ("P10", "P50", "P80", "P90"):
        blk = pct.get(key) or {}
        ld = exposure_for_date(blk.get("date"), damages)
        tc = exposure_for_delta(blk.get("workdays_vs_deterministic"), damages,
                                "workdays (target calendar, vs. deterministic)")
        _row(ws, rr, [key, blk.get("date"), ld.formula_text,
                     blk.get("workdays_vs_deterministic"), tc.formula_text])
        rr += 1
    rr += 1
    det_ld = exposure_for_date(pct.get("deterministic_engine_finish"), damages)
    rec_ld = exposure_for_date(pct.get("record_finish"), damages)
    ws.cell(row=rr, column=1, value="Deterministic engine finish LD exposure:"
            ).font = Font(bold=True, size=10)
    ws.cell(row=rr, column=2, value=det_ld.formula_text).font = Font(size=10)
    rr += 1
    ws.cell(row=rr, column=1, value="Record finish LD exposure:").font = Font(bold=True, size=10)
    ws.cell(row=rr, column=2, value=rec_ld.formula_text).font = Font(size=10)


# ---------------------------------------------------------------------------
# Disclosures
# ---------------------------------------------------------------------------
def _write_disclosures(wb, sim: dict[str, Any]):
    ws = wb.create_sheet("Disclosures")
    r = _title(ws, "Disclosures")

    def _block(title, items):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11)
        r += 1
        if not items:
            ws.cell(row=r, column=1, value="— none —").font = Font(italic=True, size=10)
            r += 1
        for it in items:
            ws.cell(row=r, column=1, value=str(it)).font = Font(size=10)
            r += 1
        r += 1

    _block("Simulation disclosures", sim.get("disclosures") or [])
    _block("Preliminary label", [sim.get("preliminary")] if sim.get("preliminary") else [])
    ws.column_dimensions["A"].width = 110


def write_sra_workbook(sim_dict: dict[str, Any], out_path: str,
                       damages: Optional[DamagesConfig] = None) -> str:
    """Write the SRA workbook: Summary, Criticality & Cruciality, Input
    Provenance, Risk Events, Sample (offsets + dates, capped at 2000 rows),
    and Disclosures.  ``sim_dict`` is the plain-dict serialization from
    ``SimulationResult.to_dict()``.

    ``damages`` (backlog S7, ANALYTICS_PROPOSAL.md §6.6) is OPTIONAL and
    strictly additive: ``None`` (the default) reproduces byte-identical
    output to before this parameter existed.  When given, the Summary
    sheet's percentile table gains time-cost and LD exposure columns (the
    classic P-date-vs-contractual-completion LD read) and the workbook
    gains an "Exposure" sheet."""
    wb = Workbook()
    _write_summary(wb, sim_dict, damages)
    _write_criticality(wb, sim_dict)
    _write_provenance(wb, sim_dict)
    _write_risk_events(wb, sim_dict)
    _write_sample(wb, sim_dict)
    if damages is not None:
        _write_exposure(wb, sim_dict, damages)
    _write_disclosures(wb, sim_dict)
    wb.save(out_path)
    return out_path
