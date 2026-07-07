"""Tests for the LI Schedule Report Card (backlog RC1-RC5).

Fixture-based assertions run against the seeded three-update series
(tests/fixtures/make_fixtures.py) already used by test_scheduleiq.py and
test_li_indices.py.  Grades on the fixtures are deterministic given the
spec and the fixture data; letters (not floats) are asserted so the test
survives cosmetic float-rounding changes without masking a real scoring
regression.
"""
import os
import subprocess
import sys

import pytest

SRC = os.path.join(os.path.dirname(__file__), "..", "src")
sys.path.insert(0, SRC)

from scheduleiq.ingest import load, load_many                      # noqa: E402
from scheduleiq.metrics.engine import evaluate, load_matrix        # noqa: E402
from scheduleiq.trend.series import analyze_series                 # noqa: E402
from scheduleiq.scorecard import (                                 # noqa: E402
    FileCard, SeriesCard, _default_points, _piecewise_score, load_spec,
    score, score_series, score_trace)
from scheduleiq.report.card_report import card_blocks               # noqa: E402
from scheduleiq.report.excel_card import write_card_workbook        # noqa: E402

FIX = os.path.join(os.path.dirname(__file__), "fixtures")
BASELINE = os.path.join(FIX, "demo_baseline.xer")
U1 = os.path.join(FIX, "demo_update1.xer")
U2 = os.path.join(FIX, "demo_update2.xer")


@pytest.fixture(scope="session", autouse=True)
def fixtures():
    if not os.path.exists(BASELINE):
        subprocess.run([sys.executable, os.path.join(FIX, "make_fixtures.py")],
                       check=True)


@pytest.fixture(scope="session")
def spec():
    return load_spec()


@pytest.fixture(scope="session")
def matrix():
    return load_matrix()


@pytest.fixture(scope="session")
def series():
    return analyze_series(load_many([BASELINE, U1, U2]))


@pytest.fixture(scope="session")
def file_cards(series):
    return [score(a) for a in series.assessments]


@pytest.fixture(scope="session")
def series_card(series):
    return score_series(series)


# ============================================================== spec loading
def test_spec_loads(spec):
    assert spec["spec_version"] == "LI-RC v1.0"
    assert spec["_sha256"]


def test_all_category_members_exist_in_matrix(spec, matrix):
    ids = {c.id for c in matrix}
    for cat in spec["categories"]:
        for cid in cat["members"]:
            assert cid in ids, f"{cat['id']} member {cid!r} is not a matrix check ID (typo?)"
    for cat in spec["series_categories"]:
        for cid in cat["members"]:
            if cid.startswith("SPEC-"):
                continue    # non-matrix synthetic member (documented in the spec header)
            assert cid in ids, f"{cat['id']} member {cid!r} is not a matrix check ID (typo?)"


def test_curve_overrides_reference_real_checks(spec, matrix):
    ids = {c.id for c in matrix}
    for cid in spec["curve_overrides"]:
        assert cid in ids, f"curve_overrides has a typo'd check id: {cid!r}"
    for cid in spec["series_curve_overrides"]:
        if cid.startswith("SPEC-"):
            continue
        assert cid in ids, f"series_curve_overrides has a typo'd check id: {cid!r}"


def test_every_critical_warning_file_check_has_exactly_one_category(spec, matrix):
    """Completeness check: every matrix check severe enough to matter for a
    single-file grade (severity critical/warning, and NOT series-only, since
    series-only checks are graded through series_categories instead) must
    belong to exactly one file-card category.  Lists any orphans/duplicates
    by ID so a future matrix edit that adds a check gets caught here."""
    member_of: dict[str, list[str]] = {}
    for cat in spec["categories"]:
        for cid in cat["members"]:
            member_of.setdefault(cid, []).append(cat["id"])

    relevant = [c for c in matrix if c.severity in ("critical", "warning")
               and c.applies_to != "series"]
    orphans = [c.id for c in relevant if len(member_of.get(c.id, [])) != 1]
    assert orphans == [], (
        f"critical/warning file-applicable checks not in exactly one "
        f"file-card category: {orphans}")

    duplicates = {cid: cats for cid, cats in member_of.items() if len(cats) > 1}
    assert duplicates == {}, f"checks claimed by more than one category: {duplicates}"


def test_series_categories_reference_defined_curves(spec):
    """Every series-category member must resolve to either a matrix-based
    default (has a real matrix threshold) or an explicit
    series_curve_overrides entry — nothing silently ungraded."""
    for cat in spec["series_categories"]:
        for cid in cat["members"]:
            assert cid in spec["series_curve_overrides"], (
                f"series category {cat['id']!r} member {cid!r} has no "
                "series_curve_overrides entry")


# ============================================================ curve arithmetic
def test_piecewise_score_known_answers():
    points = [(0, 100), (5, 70), (15, 0)]
    assert _piecewise_score(0, points) == pytest.approx(100.0)
    assert _piecewise_score(5, points) == pytest.approx(70.0)
    assert _piecewise_score(15, points) == pytest.approx(0.0)
    # midpoints, linear
    assert _piecewise_score(2.5, points) == pytest.approx(85.0)
    assert _piecewise_score(10, points) == pytest.approx(35.0)
    # clamps outside the range
    assert _piecewise_score(-5, points) == pytest.approx(100.0)
    assert _piecewise_score(50, points) == pytest.approx(0.0)


def test_default_points_max_direction(matrix, spec):
    cd = next(c for c in matrix if c.id == "DCMA-01")     # threshold 5, max
    pts = _default_points(cd, spec["curve_overrides"].get("DCMA-01"), spec)
    assert _piecewise_score(0, pts) == pytest.approx(100.0)
    assert _piecewise_score(5, pts) == pytest.approx(70.0)
    assert _piecewise_score(15, pts) == pytest.approx(0.0)   # 3x threshold


def test_default_points_min_direction(matrix, spec):
    cd = next(c for c in matrix if c.id == "DCMA-14")     # threshold 0.95, min
    pts = _default_points(cd, spec["curve_overrides"].get("DCMA-14"), spec)
    assert _piecewise_score(1.0, pts) == pytest.approx(100.0)
    assert _piecewise_score(0.95, pts) == pytest.approx(70.0)
    assert _piecewise_score(0.95 / 3, pts) == pytest.approx(0.0)


def test_zero_tolerance_collapse(matrix, spec):
    cd = next(c for c in matrix if c.id == "DCMA-09")     # threshold 0, gate
    pts = _default_points(cd, spec["curve_overrides"].get("DCMA-09"), spec)
    assert _piecewise_score(0, pts) == pytest.approx(100.0)
    assert _piecewise_score(1, pts) == pytest.approx(0.0)     # ceiling override 1


def test_log05_two_sided_band(spec):
    override = spec["curve_overrides"]["LOG-05"]
    pts = override["points"]
    assert _piecewise_score(0, pts) == pytest.approx(0.0)
    assert _piecewise_score(2, pts) == pytest.approx(100.0)
    assert _piecewise_score(4, pts) == pytest.approx(100.0)
    assert _piecewise_score(8, pts) == pytest.approx(0.0)
    assert _piecewise_score(3, pts) == pytest.approx(100.0)


# ==================================================================== score()
def test_file_cards_are_file_card_instances(file_cards):
    assert len(file_cards) == 3
    assert all(isinstance(fc, FileCard) for fc in file_cards)


def test_baseline_profile_detected_and_redistributed(file_cards):
    baseline_fc = file_cards[0]
    assert baseline_fc.profile == "baseline"
    exec_cat = next(c for c in baseline_fc.categories if c.id == "execution_performance")
    assert exec_cat.weight_used == 0.0
    # the redistributed weight lands on the other five categories, summing
    # (with resource_cost, unchanged) to 100
    total = sum(c.weight_nominal for c in baseline_fc.categories)
    assert total == pytest.approx(100.0)


def test_update_profile_weights_sum_to_100(file_cards):
    update_fc = file_cards[1]
    assert update_fc.profile == "update"
    total = sum(c.weight_nominal for c in update_fc.categories)
    assert total == pytest.approx(100.0)


def test_coverage_excludes_na(file_cards):
    for fc in file_cards:
        assert 0 <= fc.coverage_graded <= fc.coverage_total


def test_gate_status_integrity_trips_on_update1(file_cards):
    """Fixture defect: update1 seeds an actual finish after the data date
    (DCMA-09) and a status/actual-date contradiction (DAT-01) — both zero-
    tolerance gate checks — so the status-integrity gate must trip and cap
    the overall grade at 78 (a C+ ceiling) per the spec."""
    update1_fc = file_cards[1]
    gate_ids = {g.id for g in update1_fc.gates}
    assert "gate_status_integrity" in gate_ids
    assert update1_fc.overall <= 78.0
    status_cat = next(c for c in update1_fc.categories if c.id == "status_integrity")
    assert status_cat.score <= 40.0
    assert status_cat.gate_cap == 40.0


def test_gate_logic_continuity_recorded(file_cards):
    # a DCMA-12 continuity break is a documented fixture behavior (the
    # heuristic proxy trips on this synthetic network); assert the gate
    # mechanism fires and caps the category exactly at the published 60,
    # not that the underlying finding is itself "correct" schedule practice
    for fc in file_cards:
        gate = next((g for g in fc.gates if g.id == "gate_logic_continuity"), None)
        if gate is not None:
            logic_cat = next(c for c in fc.categories if c.id == "logic_network")
            assert logic_cat.score <= 60.0
            assert logic_cat.gate_cap == 60.0


def test_fixture_grades_are_stable(file_cards):
    """Deterministic on the fixtures — assert the letter, not the float, so
    a cosmetic rounding change does not mask (or hide) a real regression."""
    assert file_cards[0].letter == "C+"    # baseline
    assert file_cards[1].letter == "C"     # update1
    assert file_cards[2].letter == "D"     # update2


def test_top_factors_are_sorted_and_bounded(file_cards):
    for fc in file_cards:
        pts = [t[0] for t in fc.top_factors]
        assert pts == sorted(pts, reverse=True)
        assert sum(pts) <= 100.0 + 1e-6


# ============================================================= score_series()
def test_series_card_is_series(series_card):
    assert isinstance(series_card, SeriesCard)
    assert series_card.is_series
    assert len(series_card.file_cards) == 3


def test_series_gate_trips_on_retroactive_change(series_card):
    """Fixture defect: update2 retroactively changes 1010's previously
    reported actual start (TRD-05) — the series gate must trip."""
    gate_ids = {g.id for g in series_card.gates}
    assert "gate_record_discipline" in gate_ids
    assert series_card.overall <= 78.0
    rd_cat = next(c for c in series_card.series_categories if c.id == "record_discipline")
    assert rd_cat.score <= 40.0
    assert rd_cat.gate_cap == 40.0


def test_series_grade_is_stable(series_card):
    assert series_card.letter == "D"


def test_trajectory_present_and_bounded(series_card):
    t = series_card.trajectory
    assert t is not None
    assert t.level is not None and 0 <= t.level <= 100
    assert t.score is not None and 0 <= t.score <= 100


def test_single_file_series_card_is_file_card_only():
    sched = load(BASELINE)[0]
    a = evaluate(sched)
    from scheduleiq.trend.series import SeriesAnalysis
    sa = SeriesAnalysis(schedules=[sched], assessments=[a])
    sc = score_series(sa)
    assert sc.is_series is False
    assert len(sc.file_cards) == 1
    assert sc.series_categories == []
    assert sc.trajectory is None
    assert sc.overall == sc.file_cards[0].overall
    assert sc.letter == sc.file_cards[0].letter


# =================================================================== variant
def test_internal_variant_identical_grades_today(series):
    standard = score_series(series, variant="standard")
    internal = score_series(series, variant="internal")
    assert standard.overall == pytest.approx(internal.overall)
    assert standard.letter == internal.letter
    assert internal.internal_indices and all(i["weight"] == 0 for i in internal.internal_indices)
    assert standard.internal_indices == []


# =================================================================== tracing
def test_score_trace_self_consistent_file(file_cards):
    fc = file_cards[1]     # the one with a tripped overall-capping gate
    trace = score_trace(fc)
    wsum = sum(c["weight_used"] for c in trace["categories"] if c["score"] is not None)
    raw = sum(c["score"] * c["weight_used"] for c in trace["categories"]
             if c["score"] is not None) / wsum
    overall = raw
    if trace["overall_gate_cap"] is not None:
        overall = min(overall, trace["overall_gate_cap"])
    assert overall == pytest.approx(trace["overall"], abs=1e-6)
    assert raw == pytest.approx(trace["overall_raw"], abs=1e-6)


def test_score_trace_self_consistent_series(series_card):
    trace = score_trace(series_card)
    cats = trace["series_categories"] + [{"score": trace["trajectory"]["score"],
                                          "weight_used": trace["trajectory"]["weight"]
                                          if trace["trajectory"]["score"] is not None else 0.0}]
    wsum = sum(c["weight_used"] for c in cats if c["score"] is not None)
    raw = sum(c["score"] * c["weight_used"] for c in cats if c["score"] is not None) / wsum
    overall = raw
    if trace["overall_gate_cap"] is not None:
        overall = min(overall, trace["overall_gate_cap"])
    assert overall == pytest.approx(trace["overall"], abs=1e-6)
    assert raw == pytest.approx(trace["overall_raw"], abs=1e-6)


def test_write_trace_round_trips(series_card, tmp_path):
    from scheduleiq.scorecard import write_trace
    import json
    out = write_trace(series_card, str(tmp_path / "score_trace.json"))
    with open(out, encoding="utf-8") as f:
        data = json.load(f)
    assert data["overall"] == pytest.approx(series_card.overall)
    assert data["spec_version"] == "LI-RC v1.0"


# ================================================================ report/xlsx
def test_card_blocks_render(series_card):
    blocks = card_blocks(series_card)
    assert blocks
    assert any(b.get("type") == "h2" and "REPORT CARD" in b.get("text", "")
              for b in blocks)


def test_card_workbook_writes(series_card, tmp_path):
    path = write_card_workbook(series_card, str(tmp_path / "report_card.xlsx"))
    assert os.path.exists(path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert "Report Card" in wb.sheetnames
    assert "Category Detail" in wb.sheetnames
    assert "Score Trace" in wb.sheetnames
