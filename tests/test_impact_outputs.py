"""Tests for the milestone impact diagnostic outputs (backlog A3, ADR-0007).

Exercises the workbook writer (report/excel_impact.py), the two figures
(report/impact_figures.py), and the additive runner.py wiring: an
engine-consistent fixture (demo_impact.xer, handshake 100%) produces the
full artifact set with no refusal message, while the legacy demo_baseline /
demo_update1 / demo_update2 series (which fails engine validation) completes
with the SET-02 refusal message and NO impact workbook — every pre-existing
artifact from that series must still be produced.
"""
import os
import subprocess
import sys

import openpyxl
import pytest

SRC = os.path.join(os.path.dirname(__file__), "..", "src")
sys.path.insert(0, SRC)

from scheduleiq.ingest import load                                        # noqa: E402
from scheduleiq.analytics.impact import run_impact_analysis               # noqa: E402
from scheduleiq.analytics.asbuilt import reconstruct_asbuilt_paths        # noqa: E402
from scheduleiq.report.excel_impact import write_impact_workbook          # noqa: E402
from scheduleiq.report.impact_figures import (asbuilt_figure,             # noqa: E402
                                              waterfall_figure)
from scheduleiq.runner import run                                          # noqa: E402

FIX = os.path.join(os.path.dirname(__file__), "fixtures")
IMPACT = os.path.join(FIX, "demo_impact.xer")
BASELINE = os.path.join(FIX, "demo_baseline.xer")
U1 = os.path.join(FIX, "demo_update1.xer")
U2 = os.path.join(FIX, "demo_update2.xer")

EXPECTED_SHEETS = {"Waterfall", "Constraint Attribution", "Criticality (P5)",
                  "Calendar Restatement", "Open Ends", "As-Built Paths (P6)",
                  "Disclosures"}


@pytest.fixture(scope="session", autouse=True)
def fixtures():
    if not os.path.exists(IMPACT):
        subprocess.run([sys.executable, os.path.join(FIX, "make_fixtures.py")],
                       check=True)


@pytest.fixture(scope="session")
def sched():
    return load(IMPACT)[0]


@pytest.fixture(scope="session")
def impact_dict(sched):
    ia = run_impact_analysis(sched)
    d = ia.to_dict()
    d["data_date"] = sched.data_date.isoformat() if sched.data_date else None
    return d


@pytest.fixture(scope="session")
def asbuilt_dict(sched):
    return reconstruct_asbuilt_paths(sched).to_dict()


def _all_cell_values(ws) -> list:
    vals = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                vals.append(cell.value)
    return vals


def _wb_all_values(path) -> dict:
    wb = openpyxl.load_workbook(path)
    return {name: _all_cell_values(wb[name]) for name in wb.sheetnames}


# --------------------------------------------------------------------- workbook
def test_workbook_has_expected_sheets(impact_dict, asbuilt_dict, tmp_path):
    path = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "impact.xlsx"))
    assert os.path.exists(path)
    assert os.path.getsize(path) > 1000
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == EXPECTED_SHEETS


def test_workbook_carries_preliminary_label_on_every_sheet(impact_dict, asbuilt_dict, tmp_path):
    path = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "impact.xlsx"))
    wb = openpyxl.load_workbook(path)
    for name in wb.sheetnames:
        ws = wb[name]
        found = any(cell.value and "PRELIMINARY" in str(cell.value)
                   for row in ws.iter_rows(max_row=4) for cell in row)
        assert found, f"sheet {name!r} missing a PRELIMINARY stamp in its header rows"


def test_workbook_baseline_record_date(impact_dict, asbuilt_dict, tmp_path):
    path = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "impact.xlsx"))
    values = _wb_all_values(path)
    flat = [str(v) for v in values["Waterfall"]]
    assert any("2025-05-28" in v for v in flat)   # tool-of-record MS-IMP finish


def test_workbook_key_deltas_present(impact_dict, asbuilt_dict, tmp_path):
    path = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "impact.xlsx"))
    values = _wb_all_values(path)
    waterfall_vals = values["Waterfall"]
    for expected in (39, 2, -7):
        assert expected in waterfall_vals, f"delta {expected} not found in Waterfall sheet"


def test_workbook_p5_rows_present(impact_dict, asbuilt_dict, tmp_path):
    path = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "impact.xlsx"))
    values = _wb_all_values(path)
    crit_vals = [str(v) for v in values["Criticality (P5)"]]
    assert "ID10" in crit_vals and "manufactured" in crit_vals
    assert "MS-IMP" in crit_vals and "masked" in crit_vals


def test_workbook_asbuilt_chain_rows_present(impact_dict, asbuilt_dict, tmp_path):
    path = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "impact.xlsx"))
    values = _wb_all_values(path)
    ab_vals = [str(v) for v in values["As-Built Paths (P6)"]]
    assert "IA10" in ab_vals and "IA20" in ab_vals


def test_workbook_deterministic_on_reread(impact_dict, asbuilt_dict, tmp_path):
    p1 = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "a.xlsx"))
    p2 = write_impact_workbook(impact_dict, asbuilt_dict, str(tmp_path / "b.xlsx"))
    v1, v2 = _wb_all_values(p1), _wb_all_values(p2)
    assert v1.keys() == v2.keys()
    for name in v1:
        assert v1[name] == v2[name], f"sheet {name!r} differs between two writes"


def test_workbook_degrades_on_empty_asbuilt(impact_dict, tmp_path):
    from scheduleiq.ingest.model import Schedule
    empty_asbuilt = reconstruct_asbuilt_paths(Schedule()).to_dict()
    path = write_impact_workbook(impact_dict, empty_asbuilt, str(tmp_path / "empty.xlsx"))
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == EXPECTED_SHEETS


# ---------------------------------------------------------------------- figures
def test_waterfall_figure_nonempty(impact_dict, tmp_path):
    p = waterfall_figure(impact_dict, str(tmp_path / "waterfall.png"))
    assert os.path.exists(p)
    assert os.path.getsize(p) > 1000


def test_asbuilt_figure_nonempty(asbuilt_dict, tmp_path):
    p = asbuilt_figure(asbuilt_dict, str(tmp_path / "asbuilt.png"))
    assert os.path.exists(p)
    assert os.path.getsize(p) > 500


def test_asbuilt_figure_degrades_on_no_chains(tmp_path):
    from scheduleiq.ingest.model import Schedule
    empty = reconstruct_asbuilt_paths(Schedule()).to_dict()
    p = asbuilt_figure(empty, str(tmp_path / "asbuilt_empty.png"))
    assert os.path.exists(p)
    assert os.path.getsize(p) > 500


def test_figures_deterministic(impact_dict, tmp_path):
    p1 = waterfall_figure(impact_dict, str(tmp_path / "w1.png"))
    p2 = waterfall_figure(impact_dict, str(tmp_path / "w2.png"))
    with open(p1, "rb") as f1, open(p2, "rb") as f2:
        assert f1.read() == f2.read()


# ----------------------------------------------------------------- runner e2e
def test_runner_end_to_end_single_impact_file(tmp_path):
    out = str(tmp_path / "out_single")
    rr = run([IMPACT], out, make_pdf=False)
    assert not any("skipped" in m or "refused" in m for m in rr.messages), rr.messages
    xlsx = os.path.join(out, "impact_diagnostic.xlsx")
    wf = os.path.join(out, "fig_impact_waterfall.png")
    ab = os.path.join(out, "fig_asbuilt_paths.png")
    assert os.path.exists(xlsx) and os.path.getsize(xlsx) > 1000
    assert os.path.exists(wf) and os.path.getsize(wf) > 1000
    assert os.path.exists(ab) and os.path.getsize(ab) > 500
    assert xlsx in rr.outputs and wf in rr.outputs and ab in rr.outputs
    assert os.path.exists(os.path.join(out, "schedule_assessment.docx"))


def test_runner_end_to_end_legacy_series_refuses_gracefully(tmp_path):
    out = str(tmp_path / "out_series")
    rr = run([BASELINE, U1, U2], out, make_pdf=False)
    assert any("SET-02 handshake below threshold" in m
              and "engine impact analytics refused" in m for m in rr.messages), rr.messages
    assert not os.path.exists(os.path.join(out, "impact_diagnostic.xlsx"))
    assert not os.path.exists(os.path.join(out, "fig_impact_waterfall.png"))
    assert not os.path.exists(os.path.join(out, "fig_asbuilt_paths.png"))
    # every pre-existing artifact for a legacy series run must still be produced
    for name in ("demo_baseline_results.xlsx", "demo_update1_results.xlsx",
                "demo_update2_results.xlsx", "trend_analysis.xlsx",
                "report_card.xlsx", "score_trace.json", "path_analysis.xlsx",
                "intake_review.xlsx", "statistical_analysis.xlsx",
                "schedule_assessment.docx", "audit/audit_log.jsonl"):
        assert os.path.exists(os.path.join(out, name)), f"missing legacy artifact {name}"
