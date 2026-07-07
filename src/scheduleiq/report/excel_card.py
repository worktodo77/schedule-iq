"""RC3/RC4 — the LI Schedule Report Card as a standalone Excel workbook.

Three sheets: "Report Card" (the formatted card face — series card, if this
is a multi-file run, followed by every per-file card), "Category Detail"
(every scored member: value, score, weight, contribution — the arithmetic
that produced each category score), and "Score Trace" (a flattened
key/value rendering of the same score_trace.json this run also writes to
out_dir/score_trace.json, so an analyst who prefers Excel to JSON can still
see every intermediate number).

Reuses the LI palette/style helpers from report/excel.py rather than
redefining them (per RC3/4 instructions: import only, do not fork the
style).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from ..scorecard import FileCard, SeriesCard, score_trace
from .excel import GRAY, STATUS_FILL, TEAL, _border, _header, _row  # noqa: F401


def _grade_row(ws, r, label, value):
    ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
    ws.cell(row=r, column=2, value=value).font = Font(size=10)


def _write_card_face(ws, r: int, title: str, overall: float, letter: str,
                     spec_version: str, graded: int, total: int,
                     gates: list, categories) -> int:
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=13, color=TEAL)
    r += 1
    _grade_row(ws, r, "Overall", f"{letter}  ({overall:.0f}/100)"); r += 1
    _grade_row(ws, r, "Spec version", spec_version); r += 1
    _grade_row(ws, r, "Coverage", f"graded {graded} of {total} checks"); r += 1
    for g in gates:
        caps = "; ".join(f"{cid} capped {cap}" for cid, cap in g.category_caps.items())
        line = f"GATE TRIPPED: {g.name} ({g.rule_text}) -> {caps}"
        if g.overall_cap is not None:
            line += f"; overall capped {g.overall_cap:g}"
        ws.cell(row=r, column=1, value=line).font = Font(size=10, bold=True, color="B00000")
        r += 1
    r += 1
    _header(ws, r, ["Category", "Grade", "Score", "Weight used", "Gate cap"],
           [32, 10, 10, 12, 10])
    r += 1
    for c in categories:
        if c.score is None:
            _row(ws, r, [c.name, "—", "—", 0, "—"])
        else:
            _row(ws, r, [c.name,
                        "GATE" if c.gate_cap is not None else "",
                        round(c.score, 1), c.weight_used,
                        c.gate_cap if c.gate_cap is not None else "—"])
        r += 1
    return r + 1


def _write_top_factors(ws, r: int, top_factors: list, name_by_id: dict) -> int:
    ws.cell(row=r, column=1,
           value="TOP FACTORS (points lost of 100 · check · offenders)"
           ).font = Font(bold=True, size=11, color=TEAL)
    r += 1
    _header(ws, r, ["Points lost", "Check", "Name", "Offenders"], [12, 10, 40, 10])
    r += 1
    for pts, cid, n in top_factors:
        _row(ws, r, [round(pts, 2), cid, name_by_id.get(cid, ""), n])
        r += 1
    return r + 1


def _name_by_id(categories) -> dict:
    out = {}
    for c in categories:
        for m in c.members:
            out[m.check_id] = m.name
    return out


def _write_category_detail(ws, categories, scope: str, r: int) -> int:
    for c in categories:
        for m in c.members:
            _row(ws, r, [scope, c.name, m.check_id, m.name, m.value, m.status,
                        m.weight, m.score,
                        round(m.weight * m.score, 2) if m.score is not None else None,
                        m.offender_count, m.source],
                fill=STATUS_FILL.get(m.status))
            r += 1
    return r


def write_card_workbook(card, path: str) -> str:
    is_series = isinstance(card, SeriesCard) and card.is_series
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Card"
    r = 1
    if is_series:
        r = _write_card_face(ws, r, "LI SCHEDULE REPORT CARD — SERIES",
                             card.overall, card.letter, card.spec_version,
                             card.coverage_graded, card.coverage_total,
                             card.gates, list(card.series_categories) +
                             ([_traj_as_category(card.trajectory)] if card.trajectory else []))
        r = _write_top_factors(ws, r, card.top_factors, _name_by_id(card.series_categories))
        latest = card.file_cards[-1]
    else:
        latest = card.file_cards[-1] if isinstance(card, SeriesCard) else card

    r = _write_card_face(ws, r, f"LI SCHEDULE REPORT CARD — {latest.schedule_label}",
                         latest.overall, latest.letter, latest.spec_version,
                         latest.coverage_graded, latest.coverage_total,
                         latest.gates, latest.categories)
    r = _write_top_factors(ws, r, latest.top_factors, _name_by_id(latest.categories))
    ws.column_dimensions["A"].width = 36

    ws = wb.create_sheet("Category Detail")
    _header(ws, 1, ["Scope", "Category", "Check", "Name", "Value", "Status",
                    "Weight", "Score", "Contribution", "Offenders", "Source"],
           [10, 30, 10, 34, 10, 12, 8, 8, 12, 10, 12])
    r = 2
    if is_series:
        r = _write_category_detail(ws, card.series_categories, "Series", r)
    r = _write_category_detail(ws, latest.categories, "Latest file", r)
    ws.auto_filter.ref = f"A1:K{max(r - 1, 1)}"

    ws = wb.create_sheet("Score Trace")
    _header(ws, 1, ["Path", "Value"], [70, 40])
    rows: list = []
    _flatten(score_trace(card), "", rows)
    for i, (k, v) in enumerate(rows, 2):
        _row(ws, i, [k, v])
    ws.auto_filter.ref = f"A1:B{max(len(rows) + 1, 1)}"

    wb.save(path)
    return path


def _traj_as_category(t):
    from ..scorecard import CategoryScore
    return CategoryScore("trajectory", "File-Quality Trajectory", t.weight,
                         t.weight if t.score is not None else 0.0, t.score, t.score,
                         None, [], 0, 0)


def _flatten(obj, prefix: str, rows: list) -> None:
    if isinstance(obj, dict):
        for k, v in obj.items():
            _flatten(v, f"{prefix}.{k}" if prefix else str(k), rows)
    elif isinstance(obj, (list, tuple)):
        for i, v in enumerate(obj):
            _flatten(v, f"{prefix}[{i}]", rows)
    else:
        rows.append((prefix, obj))
